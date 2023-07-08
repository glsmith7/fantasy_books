from copy import copy
import datetime
import d20
import icons as i
from lorem_text_fantasy import lorem as lf
from math import ceil
from openpyxl.styles import Font as openpyxl_font
from openpyxl import load_workbook
from openpyxl import Workbook

import pandas as pd
import PySimpleGUI as sg
import os
import random as random
import rpg_tables as r
import shutil
import string as string
import sys
import time
import uuid
import yaml

# use faster C code for yaml if available, otherwise pure python code
try:
    from yaml import CSafeLoader as SafeLoader

except ImportError:
    from yaml import SafeLoader

# settings files
global config, master_list_stats, preferences

# logging boilerplate
import settings_GLS as s
import logging
import logging_tools_GLS
logger = logging.getLogger(__name__)

def load_settings():

    loaded_settings_files = False

    while not loaded_settings_files:
        # list of values for GUI use

        try:
            with open("fantasy_book_settings.yaml") as f:     
                config = yaml.load(f, Loader=SafeLoader)

        except PermissionError:
            sg.popup_error ("The settings file 'fantasy_books_settings.yaml' cannot be accessed. Is it open in another program?")
            sys.exit()

        except FileNotFoundError:
            sg.popup_error ("The settings file 'fantasy_books_settings.yaml' cannot be found. Has it been moved, deleted, or renamed?\n\n There is a backup copy in the folder 'default_settings_files_backup' if needed.")
            sys.exit()

        except:
            sg.popup_error ("An error has occured with settings file 'fantasy_books_settings.yaml'.")

        else:
            pass

        try:
            with open("master_books_settings.yaml") as g:     
                master_list_stats= yaml.load (g, Loader=SafeLoader)

        except PermissionError:
            sg.popup_error ("The settings file 'master_books_settings.yaml' cannot be accessed. Is it open in another program?")
            sys.exit()

        except FileNotFoundError:
            sg.popup_error ("The settings file 'master_books_settings.yaml' cannot be found. Has it been moved, deleted, or renamed?\n\n There is a backup copy in the folder 'default_settings_files_backup' if needed.")
            sys.exit()

        except:
            sg.popup_error ("An error has occured with settings file 'master_books_settings.yaml'.")

        else: 
            pass

        try:
            with open("preferences_fantasy_books.yaml") as h:     
                preferences= yaml.load (h, Loader=SafeLoader)

        except PermissionError:
            sg.popup_error ("The settings file 'preferences_fantasy_books.yaml' cannot be accessed. Is it open in another program?")
            sys.exit()

        except FileNotFoundError:
            sg.popup_error ("The settings file 'preferences_fantasy_books.yaml' cannot be found. Has it been moved, deleted, or renamed?\n\n There is a backup copy in the folder 'default_settings_files_backup' if needed.")
            sys.exit()

        except:
            sg.popup_error ("An error has occured with settings file 'preferences_fantasy_books.yaml'.")

        else:
            pass

        loaded_settings_files = True
        return config, master_list_stats, preferences

config, master_list_stats, preferences = load_settings()

# globals

global vocab_dictionary, nt, wb_source, ws_source
global window, overshoot_toggle # for GUI
global master_excel_workbook, master_excel_worksheet,master_book_pandas_table, stats

# GUI and graphics

radio_unchecked_icon = i.radio_unchecked() 
radio_checked_icon = i.radio_checked()
books_icon = i.books_icon()
excel_icon = i.excel_icon()
settings_general_icon = i.settings_general()
settings_save_icon = i.settings_save()
settings_cancel_icon = i.settings_dismiss()
settings_reset_icon = i.settings_reset()

overshoot_toggle = sg.user_settings_get_entry('-overshoot_toggle-')
radio_keys = ('-R1-', '-R2-')

# load dictionaries and pandas tables

vocab_dictionary = {}
name_tables_dictionary = config['name_SQL_tables']
nt={}

for key,table in name_tables_dictionary.items():
    nt[key] = r.RPG_table(table)

nt['complete_table_female_names'].description = "Female Names Amalgamated"
nt['complete_table_male_names'].description = "Male Names Amalgamated"
nt['titles_saints_amalgamated']= nt['titles_saints_male'] + nt['titles_saints_female']
nt['titles_person_famous_amalgamated'] = nt['titles_person_famous_male'] + nt['titles_person_famous_female']

# load first name table dictionaries

nt['name_tables_male'] = {}
nt['name_tables_female'] = {}
nt['surnames_tables'] = {}

for i in config['list_of_names_tables_male']:
    nt['name_tables_male'][i] = r.RPG_table(i)
    nt['complete_table_male_names'] = (nt['name_tables_male'][i]) + nt['complete_table_male_names']

for i in config['list_of_names_tables_female']:
    nt['name_tables_female'][i] = r.RPG_table(i)
    nt['complete_table_female_names'] = (nt['name_tables_female'][i]) + nt['complete_table_female_names']

# load surnames table dictionaries

for i in config['list_of_surnames_tables']:
    nt['surnames_tables'][i] = r.RPG_table(i) # creates dictionary containing a table for each nationality.

######################## FUNCTIONS ########################

def about_window_gui():

    credits_text = r"by GLS, (C) 2023 \
     \
    this is a test."
    row_1 = [
        sg.Multiline(default_text=credits_text,)
        
    ]
    
    row_final = [
        sg.Push(),
        sg.Button('Ok', key='-CLOSE-ABOUT-'),
    
    ]
    layout = [
        [row_1],
        #
        [row_final]
            ]
    
    return layout

def archive_to_master(source="books_spreadsheet_out.xlsx", source_worksheet = "Book Hoard",destination="master_fantasy_book_list.xlsx",destination_worksheet = "Master List"):
    '''
    Places books in an excel spreadsheet into the master_book_list. This is all books that exist in a campaign, and is used to produce additional copies (if they are extant) of already-described books. This happens at the appropriate frequency for the total number of books in the game world.

    source = an Excel (*.xlsx) file. Defaults to "books_spreadsheet_out.xlsx".
    worksheet = the "tab" from the source Excel file. Defaults to "Book Hoard".

    destination = an Excel (*.xlsx) file. Defaults to "master_fantasy_book_list.xlsx".
    destination_worksheet = the "tab" from the destination Excel file. Defaults to "Master List"

    '''
    # load source
    try:
        wb_source_books = load_workbook(filename= source)
    except FileNotFoundError:

        raise FileNotFoundError ("Could not load the source file: " + source + ".")

    if source_worksheet in wb_source_books.sheetnames: 
        ws_source_books = wb_source_books[source_worksheet]
    else:
        raise FileNotFoundError ("Could not find the worksheet: " + source_worksheet + " even though file " + source + " was successfully loaded.")
    
    # load destination

    try:
        wb_dest = load_workbook(filename= destination)
    except:
        create_new_master_excel_file(filename=destination)
    wb_dest = load_workbook(filename= destination)

    if destination_worksheet in wb_dest.sheetnames: 
        ws_dest = wb_dest[destination_worksheet]
    else:
        ws_dest = wb_dest.create_sheet(title=destination_worksheet)
        book_columns,current_language_index, flavor_title_index = book_characteristics(books)

        # column headers to new sheet
        the_counter = 0
        for item in book_columns:
            the_counter += 1
            ws_dest.cell(row=1,column=the_counter,value=item)
            ws_dest.cell(row=1,column=the_counter).font = openpyxl_font(bold='bold',size=9)

    # Make sure can save

    try_to_save = True
    while try_to_save:
        try:
            wb_dest.save(destination) 
        except PermissionError:
            user_response = sg.popup_ok_cancel ("The Excel file " + destination + " can't be opened, which means it is probably open in another program. \n\n Close the file, and click OK to try again.\n\nTo quit without saving, click CANCEL.")
            
            if user_response == "Cancel":
                try_to_save = False
                sg.popup_ok ("Quitting without saving to Excel.")
                sys.exit()
                break
        else:
            try_to_save = False # ie succeeded
    ######
    sg.theme("Dark Green 5")
    row_dest = ws_dest.max_row + 1
    the_count = 0

    # copy each cell from source to destination
    total_number_to_copy = (ws_source_books.max_row) - (ws_source_books.min_row)
    for row_source in ws_source_books.iter_rows(min_row=ws_source_books.min_row+1, min_col=ws_source_books.min_column, max_row=ws_source_books.max_row, max_col=ws_source_books.max_column):
        
        the_count +=1 
        the_note = row_source[config['NOTE_COLUMN_INDEX']]

        if not sg.one_line_progress_meter(
                'Copy to master', 
                the_count, 
                total_number_to_copy, 
                orientation = 'h',
                ) and the_count+1 != total_number_to_copy:
                sg.popup_auto_close("The rest of the books won't be copied to the master library Excel file.")
                break
        
        if "do_not_archive" == the_note.value or "has_been_archived" == the_note.value:
            continue

        else:

            the_note.value = "has_been_archived"

        for cell_source in row_source:
            dest_coords = str(cell_source.column_letter) + str(row_dest)
            cell_dest = ws_dest[dest_coords]
            
            cell_dest.value = cell_source.value
            cell_dest.font = copy (cell_source.font)
        
        row_dest += 1

    sg.popup_notify("Finished transfer to master file \n\n" + str (destination) + "\n\nin worksheet\n\n" + str (destination_worksheet) + ".",
                    title = "Finished!",
                    display_duration_in_ms = config['duration_toaster_popups'],
                    fade_in_duration = config['fade_in_duration_toaster_popups'],
                    alpha = config['alpha_toaster_popups'],
                    location = None)
    
    wb_source_books.save(source)
    wb_dest.save(destination)
    
    wb_source_books.close()
    wb_dest.close()

def backup_excel_file(filename = "master_fantasy_book_list.xlsx"):

    # archive the old master file
    file_source = filename
    the_label = datetime.datetime.now().strftime("%Y%m%d_%H%M_%S")
    file_destination = 'excel_backups\master_fantasy_book_list_backup_XXLABELXX.xlsx'
    file_destination = file_destination.replace('XXLABELXX',the_label)
    
    try_to_save = True
    while try_to_save:
        try:
            shutil.move(file_source, file_destination)

        except PermissionError:
            
            user_response = sg.popup_ok_cancel ("The Excel file " + file_source + " can't be opened, which means it is probably open in another program. \n\n Close the file, and click OK to try again.\n\nTo quit without saving, click CANCEL.")
            
            if user_response == "Cancel":
                try_to_save = False
                sg.popup_ok ("Quitting without saving to Excel.")
                sys.exit()
                break

        except FileNotFoundError:
            sg.popup_notify('The old file ' + file_source + ' could not be archived, probably because it has been either renamed or does not exist. Try creating an empty excel file named ' + file_source,
                    title = "File not found",
                    icon = '', # TO_DO
                    display_duration_in_ms = 5000,
                    fade_in_duration = config['fade_in_duration_toaster_popups'],
                    alpha = 1,
                    location = None)
        
        except:
            raise ("A new file could not be created.")
        
        else:
            try_to_save = False # ie succeeded
            sg.popup_notify('Former master book file moved to folder "excel_backups".',
                        title = "Archived",
                        icon = '', # TO_DO
                        display_duration_in_ms = config['duration_toaster_popups'],
                        fade_in_duration = config['fade_in_duration_toaster_popups'],
                        alpha = config['alpha_toaster_popups'],
                        location = None)
    
def book_characteristics(books):


    book_attributes = [attribute for attribute in dir(books[1])
                   if not attribute.startswith('__')
                   and not callable(getattr(books[1], attribute))
                   ]
    
        # this bit adds any variables that have been omitted from the above list, so all will be displayed even if user error/omission.
    for item in book_attributes:
        if item not in config['book_variables_in_chosen_order']:
            config['book_variables_in_chosen_order'].append(item)

    current_language_index = config['book_variables_in_chosen_order'].index('current_language')
    flavor_title_index = config['book_variables_in_chosen_order'].index('book_title_flavor')

    return config['book_variables_in_chosen_order'], current_language_index+1, flavor_title_index+1 # index starts 0, Excel starts 1

def book_batch (number=1, **kwargs):
    
    ''' 
    Produces a given number of books. Randomized characteristics unless keyword parameters are passed in. Those not passed with be randomized as far as it able (some values are interrelated, and so this can result in some slight deviations from the tables.)
    '''
    
    books = {}
    
    sg.theme('Light Blue 1')

    while books == {}:
        
        running_total = 0
        the_count = 0

        while the_count < number:

            the_count += 1
            if check_if_should_place_existing_title():
                books[the_count] = pick_existing_book()

            else:
                books[the_count] = create_fantasy_book(**kwargs)

            running_total += books[the_count].market_value
            
            if not sg.one_line_progress_meter(
                'Generating books', 
                the_count, 
                number, 
                orientation = 'h',
                ):
                sg.popup_auto_close("The rest of the books won't be generated. Those already made will be added to the Excel file and the master library Excel file.")
                break

    sg.one_line_progress_meter_cancel()
    return books, running_total

def book_hoard (value_of_books=0,overshoot=True, **kwargs):
    ''' 
    produces a list of books worth the passed value. If overshoot is False, keeps total worth equal to or under value. If overshoot is true, then will produce a list that is _at least_ the passed value.

    Randomized characteristics unless keyword parameters are passed in. Those not passed with be randomized as far as it able (some values are interrelated, and so this can result in some slight deviations from the tables.)
    '''

    books = {}
    sg.theme('Light Blue 1')

    while books == {}:
    
        running_total = 0
        the_count = 0
        
        while running_total < value_of_books:
            the_count += 1
            if check_if_should_place_existing_title():
                books[the_count] = pick_existing_book()

            else:
                books[the_count] = create_fantasy_book(**kwargs)

            running_total += books[the_count].market_value

            if not sg.one_line_progress_meter(
                'Generating books', 
                running_total, 
                value_of_books, 
                orientation = 'h',
                
                ):
                sg.popup_auto_close("The rest of the books won't be generated. Those already made will be added to the Excel file and the master library Excel file.")
                break

        sg.one_line_progress_meter_cancel()
        if overshoot: 
            pass    
        else:
            running_total -= books[len(books)].market_value # subtract last value that put us over the top
            books.popitem() # delete last book which put over the top

        if books == {}:
            sg.popup_notify("Zero books made in hoard; retrying ....",
                    title = "Retrying",
                    icon = radio_unchecked_icon,
                    display_duration_in_ms = config['duration_toaster_popups'],
                    fade_in_duration = config['fade_in_duration_toaster_popups'],
                    alpha = config['alpha_toaster_popups'],
                    location = None)
    
    return books, running_total

def calculate_stats_excel (wb_master,ws_master):
    
    col_list = ['market_value','number_extant_copies','number_extant_available_to_place']
    total={}
    total['rows'] = ws_master.max_row-1 # total rows less 1 heading.

    for column in col_list:
        
        total[column] = 0 # will be non-existence if Excel file is empty, i.e., less than 2 lines, see for m in range loop.
        for m in range(2,ws_master.max_row+1): # starts from 2 since line 1 is headers. Max row thus needs +1 as well.
           # the_column_index_num = 
            total[column] = total[column] + int (ws_master.cell(row=m,column=config[column]).value)
            
    return (total)

def check_radio(key): # GUI function
    radio_keys = ('-R1-', '-R2-')
    
    for k in radio_keys:
        window1[k].update(radio_unchecked_icon)
        window1[k].metadata = False
    window1[key].update(radio_checked_icon)
    window1[key].metadata = True

def check_if_should_place_existing_title():

    if stats['number_extant_available_to_place'] < 1: # ie none exist to place
        return False
    
    else: #TO_DO

        total_books_copies_in_campaign = preferences['TOTAL_BOOKS_IN_CAMPAIGN']
        total_books_copies_discovered = stats['number_extant_available_to_place']
        dice_string = "1d" + str (total_books_copies_in_campaign)
        the_roll = d20.roll(dice_string).total
        if the_roll <= total_books_copies_discovered:
            return True
        
        else:
            return False

def create_fantasy_book(book_type=None, **kwargs):
    ''' Returns a book object. Type can be default (normal), esoteric, authority, or magic'''
    book_type = string.capwords(str(book_type))
    return FantasyBook(**kwargs)

def create_new_master_excel_file(filename = 'master_fantasy_book_list.xlsx'):
    file_source = 'blank_excel_files_templates\master_fantasy_book_list_BLANK.xlsx'
    file_destination = filename
    if file_destination[-5:] != ".xlsx":
        file_destination = file_destination + ".xlsx"

    try_to_save = True
    while try_to_save:
        try:
            shutil.copyfile(file_source, file_destination)

        except PermissionError:
            
            user_response = sg.popup_ok_cancel ("The Excel file " + file_source + " can't be opened, which means it is probably open in another program. \n\n Close the file, and click OK to try again.\n\nTo quit without saving, click CANCEL.")
            
            if user_response == "Cancel":
                try_to_save = False
                sg.popup_ok ("Quitting without saving to Excel.")
                sys.exit()
                break

        except FileNotFoundError:
            sg.popup_notify('The template file ' + file_source + ' could not be found, probably because it has been either renamed or does not exist. Try creating an empty excel file named ' + file_source,
                    title = "File not found",
                    icon = '', # TO_DO
                    display_duration_in_ms = 5000,
                    fade_in_duration = config['fade_in_duration_toaster_popups'],
                    alpha = 1,
                    location = None)
        
        except:
            raise ("A new file could not be created.")
        
        else:
            try_to_save = False # ie succeeded
            sg.popup_notify('New blank master file created.',
                    title = "New master",
                    icon = radio_unchecked_icon,
                    display_duration_in_ms = config['duration_toaster_popups'],
                    fade_in_duration = config['fade_in_duration_toaster_popups'],
                    alpha = config['alpha_toaster_popups'],
                    location = None)

def export_books_to_excel (books,filename = 'books_spreadsheet_out.xlsx', worksheet = 'Book Hoard'):
    
    '''
    Takes a list of books and exports to Excel file. Default file 'books_spreadsheet_out.xlsx' in same folder, worksheet defaults to "Book Hoard."
    filename = 
    worksheet = 

    Can be passed in.

    If file and worksheet exist, is appended to. If file and/or worksheet do not exist, they are created.

    Note that export cannot function if the desired file is open. If this happens an option will be presented allowing user to close the Excel file and retry the save.
    '''
    book_columns,current_language_index, flavor_title_index = book_characteristics(books)
    
    config['CURRENT_LANGUAGE_COLUMN_INDEX'] = current_language_index
    config['FLAVOR_TITLE_COLUMN_INDEX'] = flavor_title_index
    
    try:
        wb = load_workbook(filename= filename)
    except:
        wb = Workbook()

    if worksheet in wb.sheetnames: 
        ws = wb[worksheet]
    else:
        ws = wb.create_sheet(title=worksheet)

    # column headers
    the_counter = 0
    for item in book_columns:
        the_counter += 1
        ws.cell(row=1,column=the_counter,value=item)
        ws.cell(row=1,column=the_counter).font = openpyxl_font(bold='bold',size=9)

    # Make sure can save

    try_to_save = True
    while try_to_save:
        try:
            wb.save(filename) 
        except PermissionError:
            user_response = sg.popup_ok_cancel ("The Excel file " + filename + " is probably open. \n\n Close the file, and click OK to try again.\n\nTo quit without saving, click CANCEL.")
            
            if user_response == "Cancel":
                try_to_save = False
                sg.popup_ok ("Quitting without saving to Excel.")
                sys.exit()
                break
        else:
            try_to_save = False # ie succeeded

    # each row for a book
    the_counter = 0
    sg.theme("Dark Green 4")
    for book in books:
        row = []
        for attribute in book_columns:
            row.append(getattr(books[book],attribute))
        ws.append(row)
        the_counter += 1

            # now get language of the last row (just added) and set the proper font for the flavor title cell
        the_lang = ws.cell(row=ws.max_row,column=current_language_index)
        the_flavor = ws.cell(row=ws.max_row, column=flavor_title_index)
        the_flavor.font = openpyxl_font(name=config['font_languages'][the_lang.value],size=config['DEFAULT_EXCEL_FLAVOR_FONT_SIZE'])
        if not sg.one_line_progress_meter(
                'Placing books in Excel file.', 
                the_counter, 
                len(books), 
                orientation = 'h',
                ):
                sg.popup_auto_close('Saving the rest of the books to the Excel file canceled. Those books already in the file will still be entered into the master library file.')
                break
    wb.save(filename)
    wb.close()

    sg.popup_notify("Export to Excel file \n\n" + filename + "\n\nis complete.",
        title = "Excel export done.",
        icon = excel_icon,
        display_duration_in_ms = config['duration_toaster_popups'],
        fade_in_duration = config['fade_in_duration_toaster_popups'],
        alpha = config['alpha_toaster_popups'],
        location = None)
    
def fantasy_books_main_gui():
    layout = [

                # default spreadsheets x 2
            # line 1 ##################################################################
            [
             sg.Text('Default spreadsheet out:'),  
             sg.Combo(sorted(sg.user_settings_get_entry('-default_out_filenames-', [])), 
                      default_value=sg.user_settings_get_entry('-last_default_out_filename-', ''), 
                      size=(50, 1), 
                      key='-EXCEL_OUT_FILENAME-',
                      ),   
            
            sg.FileBrowse(),   
            sg.Button('Clear History',  
                 key = 'Clear_History_Default_Out',
                     ), 
            sg.Text('Worksheet:'), 
            sg.Combo(sorted(sg.user_settings_get_entry('-default_out_worksheets-', [])), 
                default_value=sg.user_settings_get_entry('-last_default_out_worksheet-', ''), 
                size=(20, 1), 
                key='-EXCEL_OUT_WORKSHEET-',
                ), 

            sg.Button('Clear History',  
                 key = 'Clear_History_Default_Out_Worksheet')
            ],
            
            # line 2 ##################################################################
            [
             sg.Text('Default master list:'),  
             sg.Combo(sorted(sg.user_settings_get_entry('-default_master_filenames-', [])), 
                    default_value=sg.user_settings_get_entry('-last_default_master_filename-', ''), 
                    size=(50, 1), 
                    key='-MASTER_FILENAME-',
                    expand_x=True,
                    disabled=True,
                    ),   
            sg.FileBrowse(
                disabled = True,
            ),   
            sg.Button('Clear History', 
                 key = 'Clear_Master_History',
                 disabled = True,
                 ),
            sg.Text('Worksheet:'), 
            sg.Combo(sorted(sg.user_settings_get_entry('-default_master_worksheets-', [])), 
                    default_value=sg.user_settings_get_entry('-last_default_master_worksheet-', ''),
                    size=(20, 1), 
                    key='-MASTER_WORKSHEET-',
                    disabled = True,
                    ), 
            sg.Button('Clear History',  
                 key = 'Clear_History_Master_Worksheet',
                 disabled = True,
                 )
            ],

            # line 3 ##################################################################
            [
                sg.Image(radio_checked_icon if sg.user_settings_get_entry('-R1_status-') else radio_unchecked_icon,
                      enable_events=True, 
                      k='-R1-', 
                      metadata=sg.user_settings_get_entry('-R1_status-'),
                      tooltip = ' A book collection of a given value will be generated. '
                      ),
            
            sg.Text('Generate books by value ⟶', 
                    enable_events=True, 
                    k='-T1-',
                    
                    ),
            
            sg.Input(
                    key = "-value_of_books_to_make-",
                    default_text = sg.user_settings_get_entry('-books_value-'),
                    size = (15, 1),
                    enable_events = True,
                    
            ),

            sg.Text('gp total',
                    expand_x=True,
                    ), 

            sg.Text('Allow last book to exceed budget:',
                    ), 

            sg.Button (
                    key='Overshoot', 
                    button_text='Yes' if sg.user_settings_get_entry('-overshoot_toggle-') else 'No', 
                    button_color='white on green' if sg.user_settings_get_entry('-overshoot_toggle-') else 'white on red',
                    size=(4, 1), 
                    tooltip=' If YES, the last book allowed to bring the hoard total to more than requested. If NO, the last book will not be included, and the hoard total value will thus be less than requested amount. ',
                    ),
            ],

            # line 4 ###################################################################
            [
                sg.Image(radio_checked_icon if sg.user_settings_get_entry('-R2_status-') else radio_unchecked_icon,
                      enable_events=True, 
                      k='-R2-', 
                      metadata=sg.user_settings_get_entry('-R2_status-'),
                      tooltip = ' A given number of books will be generated. '),

            sg.Text('Generate books by number ⟶', 
                    enable_events=True, 
                    k='-T2-',
                    tooltip = ' A given number of books will be generated. ',
                    ),
            
            sg.Input(
                    key = "-number_of_books_to_make-",
                    default_text = sg.user_settings_get_entry('-books_number-'),
                    size = (13, 1),
                    enable_events = True,
                    
            ),

            sg.Text('books',
                    expand_x=True,
                    ), 

            sg.Push(),
            sg.Button('Master library stats', tooltip = "Info about the current master library. TO IMPLEMENT", disabled=True),
            ],

            # Final buttons
            [
                sg.Button('Generate Books', 
                       bind_return_key=True,
                       ),
            sg.Button("Save settings and Quit"),  
            sg.Button('Quit'),
            sg.Push(),
            sg.Button('Reset to defaults'),
            sg.Button('Edit preferences'),
            sg.Button('Clear master Excel file'),
            ]
            ]
    
    return layout

def get_proper_random_book ():
    '''
    Picking a row at random isn't a true randomization, since each row has a different number of extant books. This routine uses a weighted random choice and returns the line.
    '''
    
    number_of_lines = len (master_book_pandas_table.index)
    lines_list = [i for i in range (0,number_of_lines)]
    weighted_chances_list = master_book_pandas_table["number_extant_available_to_place"].values.tolist()

    row_target = random.choices(lines_list,weights=weighted_chances_list,k=1)
    to_return = int (row_target[0])
    to_return += 2 # Excel starts from 1, pandas from 0. Pandas also does not have column headers here, so + 2 net to match Excel.

    return to_return

def import_language_words():
    ''' creates a dictionary of lists of various languages/character sets for the 'flavor text' titles of books based on their language.
        titles are generated with a lorem_ipsum algorithm from random words in *.txt files in the folder lorem_ipsum_fantasy.
        This is called just once as the program starts, and then the lists are passed to the lorem_ipsum_fantasy package.

        More languages and the like can be added in the config['dictionary_languages'] dictionary at the beginning of the program with the other constants. Key is the language; value is the name of the text file.
    '''

    ROOT_DIR = os.getcwd() # os.path.realpath(os.path.join(os.path.dirname(__file__), '..'))
    THIS_FOLDER = os.path.join((ROOT_DIR), 'lorem_text_fantasy')
    vocab_dictionary = {}

    for language,file in config['dictionary_languages'].items():
        TARGET_LANGUAGE_FILE = os.path.join(THIS_FOLDER, file)
        the_words_imported = []
        with open(TARGET_LANGUAGE_FILE,encoding = 'utf8', mode='r') as f:
            for line in f.readlines():
                the_words_imported.append(line.strip())
            vocab_dictionary[language] = the_words_imported

    return vocab_dictionary

def load_excel_objects (filename = 'master_fantasy_book_list.xlsx', worksheet = 'Master List'):

# load source
    try:
        wb_source = load_workbook(filename= filename)
    except:
        raise FileNotFoundError ("Could not load the source file: " + filename + ".")

    if worksheet in wb_source.sheetnames: 
        ws_source = wb_source[worksheet]
    else:
        raise FileNotFoundError ("Could not find the worksheet: " + worksheet + " even though file " + filename + " was successfully loaded.")
    
    return wb_source,ws_source

def overshoot_event(overshoot_toggle): # gui

    overshoot_toggle = not overshoot_toggle
    window1['Overshoot'].update(
        text='Yes' if overshoot_toggle else 'No', 
        button_color='white on green' if overshoot_toggle else 'white on red'
        )
    
    return overshoot_toggle
    
def pick_existing_book():
    '''
    Randomly picks a single book from the master Excel file, and passes it back as a dictionary book_to_be, which can then be treated as input for create_fantasy_book(book_to_be**)
    '''
    
    book_to_be = {} 
    
    while True:
        random_book = get_proper_random_book()

        index = config['number_extant_available_to_place']
        try:
            number_books_left_this_title = int (master_excel_worksheet.cell(row = random_book, column = index).value)
        
        except:

            break # errors if nothing in the cell, so nothing to pick; get out of loop.

        if number_books_left_this_title == 0:
            sg.popup_notify("Zero books of this title remain for placement; picking another book...",
                title = "None of these left.",
                icon = '', # TO_DO
                display_duration_in_ms = config['duration_toaster_popups'],
                fade_in_duration = config['fade_in_duration_toaster_popups'],
                alpha = config['alpha_toaster_popups'],
                location = None)
            
            random_book+=1 # ie not avail, pick another
            
            number_books_left_this_title = int (master_excel_worksheet.cell(row = random_book, column = index).value)

        # Otherwise, copy over
        master_excel_worksheet.cell(row = random_book, column = index, value = (number_books_left_this_title-1))
        
        # also edit the pandas version
        master_book_pandas_table.at[random_book-2,'number_extant_available_to_place'] = (number_books_left_this_title-1)

        the_counter = 1 # Excel columns start at 1, not zero.
        for attribute in config['book_variables_in_chosen_order']:
            book_to_be [attribute] = master_excel_worksheet.cell(row=random_book, column = the_counter).value
            the_counter += 1
                    
        break
    
    book = create_fantasy_book(**book_to_be)
    stats = calculate_stats_excel(master_excel_workbook,master_excel_worksheet) # filename = filename, worksheet=worksheet)
    update_master_books_array(stats)

    # sg.popup_notify("Pre-existing book from master library placed.",
    #                 title = "Another copy!",
    #                 icon = radio_unchecked_icon,
    #                 display_duration_in_ms = config['duration_toaster_popups'],
    #                 fade_in_duration = config['fade_in_duration_toaster_popups'],
    #                 alpha = config['alpha_toaster_popups'],
    #                 location = None)
    return book

def progress_window_gui():
    pass
    #     layout = [

                
    #             [sg.Push(),
    #              sg.Text(
    #                 text='Generating books:',
    #                 font = 'Helvetica 12 bold',

                    
    #                 ),  
    #              sg.Push(),
                
    #             ],
    #             [sg.Push(),
    #              sg.Text(
    #                 key = '-books_count-',
    #                 text='5',
    #                 font = 'Helvetica 48',
    #                 justification='Center'
    #                 ),  
    #             sg.Push(),
    #             ],
    #             [sg.Push(),
    #              sg.Text(
    #                 text='Value:',
    #                 font='Helvetica 12',
    #                 pad = (0,1),
    #             ),
    #             sg.Text(
    #                 key = '-gold_pieces-',
    #                 text=10000,
    #                 font='Helvetica 12',
    #                 pad = (0,0),
    #             ),
    #             sg.Text(
    #                 text='gp',
    #                 font='Helvetica 12',
    #                 pad = (0,1),
    #             ),
    #             sg.Push(),
    #             ],
    #             [
    #             sg.Push(),
    #             sg.ProgressBar(
    #                 key='-book_generation_progress',
    #                 style='default',
    #                 orientation = 'horizontal',
    #                 # expand_x = True,
    #                 max_value = 100,
                    
    #             ),
    #             sg.Push(),
    #             ],
    #     ]

    #     return layout

def radio_is_checked(key): # GUI
        return window1[key].metadata

def read_excel_file_into_pandas (filename = 'master_fantasy_book_list.xlsx',worksheet = 'Master List'): 
    
    try_to_save = True
    while try_to_save:
        try:
            excel_file_pandas = pd.read_excel(filename, sheet_name=worksheet, header=0, index_col=None, usecols=None, dtype=None, engine="openpyxl", decimal='.')

        except PermissionError:
                
            user_response = sg.popup_ok_cancel ("The Excel file " + filename + " can't be opened, which means it is probably open in another program. \n\n Close the file, and click OK to try again.\n\nTo quit without saving, click CANCEL.")
            
            if user_response == "Cancel":
                try_to_save = False
                sg.popup_ok ("Quitting ...")
                sys.exit()
                break

        except FileNotFoundError:
            sg.popup_notify('The old file ' + filename + ' could not be archived, probably because it has been either renamed or does not exist.',
                    title = "File not found",
                    icon = '', # TO_DO
                    display_duration_in_ms = 5000,
                    fade_in_duration = config['fade_in_duration_toaster_popups'],
                    alpha = 1,
                    location = None)
                
        except:
            sg.popup_error("Problem opening file:" + filename + " If you get a ZipFile error, this may be due to a corrupt Excel file.")
            
        else:
            try_to_save = False # succeeded!     
            return excel_file_pandas

def save_gui_settings():
    # Save combo boxes and contents - out
    sg.user_settings_set_entry('-default_out_filenames-', list(set(sg.user_settings_get_entry('-default_out_filenames-', []) + [values['-EXCEL_OUT_FILENAME-'], ])))
    sg.user_settings_set_entry('-last_default_out_filename-', values['-EXCEL_OUT_FILENAME-'])
    sg.user_settings_set_entry('-default_out_worksheets-', list(set(sg.user_settings_get_entry('-default_out_worksheets-', []) + [values['-EXCEL_OUT_WORKSHEET-'], ])))
    sg.user_settings_set_entry('-last_default_out_worksheet-', values['-EXCEL_OUT_WORKSHEET-'])

    # Save combo boxes and contents - master
    sg.user_settings_set_entry('-default_master_filenames-', list(set(sg.user_settings_get_entry('-default_master_filenames-', []) + [values['-MASTER_FILENAME-'], ])))
    sg.user_settings_set_entry('-last_default_master_filename-', values['-MASTER_FILENAME-'])
    sg.user_settings_set_entry('-default_master_worksheets-', list(set(sg.user_settings_get_entry('-default_master_worksheets-', []) + [values['-MASTER_WORKSHEET-'], ])))
    sg.user_settings_set_entry('-last_default_master_worksheet-', values['-MASTER_WORKSHEET-'])

    # sg.user_settings_set_entry('-overshoot_toggle-', overshoot_toggle)
    sg.user_settings_set_entry('-overshoot_toggle-', overshoot_toggle)

    sg.user_settings_set_entry('-books_value-', values['-value_of_books_to_make-'])
    sg.user_settings_set_entry('-books_number-', values['-number_of_books_to_make-'])

    sg.user_settings_set_entry('-R1_status-', window1["-R1-"].metadata)
    sg.user_settings_set_entry('-R2_status-', window1["-R2-"].metadata)

def save_master_books_settings():
    '''
    Saves the master_list_stats array so data persists between sessions.
    '''
    try_to_save = True
    while try_to_save:
        try:
            with open("master_books_settings.yaml", "w") as f:     
                yaml.dump(master_list_stats, stream=f, default_flow_style=False, sort_keys=False)

        except PermissionError:
            
            user_response = sg.popup_ok_cancel ("The Excel file " + 'master_books_settings.yaml' + " can't be saved, which means it is probably open in another program. \n\n Close the file, and click OK to try again.\n\nTo quit without saving, click CANCEL.")
            
            if user_response == "Cancel":
                try_to_save = False
                sg.popup_ok ("Quitting without saving to Excel.")
                sys.exit()
                break

        except FileNotFoundError: # this will probably never be called. Here in case, for now.
            sg.popup_notify('The file ' + 'master_books_settings.yaml' + ' could not be saved, probably because it has been either renamed or does not exist.',
                    title = "File not found",
                    icon = '', # TO_DO
                    display_duration_in_ms = 5000,
                    fade_in_duration = config['fade_in_duration_toaster_popups'],
                    alpha = 1,
                    location = None)
        
        except:
            raise ("A problem with saving the master library settings file was encountered.")
        
        else:
            try_to_save = False # ie succeeded

def settings_gui():

    row_1 = [
        sg.Text("Ancient languages not to be translated into:", tooltip='These are languages which are ancient or dead. \nThus, original works may be in these languages, \nand there may be translations _into_ other languages, \nbut they will not be the destination language for translations. \n(E.g., a Latin work could be translated into English, \nor could remain in Latin. An English text, \nhowever, would not be translated into Latin. )'),
        sg.Input(preferences['ANCIENT_LANGUAGES_WHICH_WOULD_NOT_BE_TRANSLATED_INTO'],key='ANCIENT_LANGUAGES_WHICH_WOULD_NOT_BE_TRANSLATED_INTO')
    ]
    row_2 = [
        sg.Text("Additional age for translation:", tooltip= 'Dice formula for age added to book that is a translation.',),
        sg.Input(preferences['TRANSLATION_ADDITIONAL_AGE_OF_ORIGINAL'], size=(8,1), key='TRANSLATION_ADDITIONAL_AGE_OF_ORIGINAL'),
        sg.Push(),
        sg.Text("Chance of being translation (%):", tooltip='Chance from 0-100% book is a translation.',size=(25,1)),
        sg.Input(preferences['CHANCE_OF_BEING_TRANSLATION'], size=(3,1),key='CHANCE_OF_BEING_TRANSLATION'),
        
    ]
    row_3 = [     
        sg.Text("Chance of female author (%):", tooltip= 'Chance from 0-100% of author having female name.',size=(25,1)),
        sg.Input(preferences['CHANCE_OF_FEMALE_AUTHOR'], size=(3,1),key='CHANCE_OF_FEMALE_AUTHOR'),
        sg.Push(),
        sg.Text("Chance of being incomplete (%):", tooltip='Chance from 0-100% book is not entirely intact.'),
        sg.Input(preferences['CHANCE_OF_INCOMPLETE_WORK'], size=(3,1),key='CHANCE_OF_INCOMPLETE_WORK'),
    ]
    row_4 = [     
        sg.Text("Chance of author title (%):", tooltip= 'Chance from 0-100% of author name including\n a title (e.g., "Doctor, Professor").',size=(25,1)),
        sg.Input(preferences['CHANCE_OF_TITLE_IN_AUTHOR_NAME'], size=(3,1),key='CHANCE_OF_TITLE_IN_AUTHOR_NAME'),
        sg.Push(),
        sg.Text("Chance of author epithet (%):", tooltip='Chance from 0-100% of author name including\n an epithet (e.g., "Bob the Brave, Joe the Fat").'),
        sg.Input(preferences['CHANCE_OF_EPITHET_IN_AUTHOR_NAME'], size=(3,1),key='CHANCE_OF_EPITHET_IN_AUTHOR_NAME'),
    ]
    row_5 = [
        sg.Text("Minimum age book:", tooltip= 'Book will not be younger than this.',size=(25,1)),
        sg.Input(preferences['MINIMUM_AGE_BOOK'], size=(3,1),key='MINIMUM_AGE_BOOK'),
        sg.Push(),
        sg.Text("Maximum age book:", tooltip='Book will not be older than this.'),
        sg.Input(preferences['MAXIMUM_AGE_BOOK'], size=(4,1),key='MAXIMUM_AGE_BOOK'),
    ]
    row_6 = [
        sg.Text("Total number of books in campaign:", tooltip= 'Total number of volumes (includes duplicates of same text).',size=(25,1)),
        sg.Input(preferences['TOTAL_BOOKS_IN_CAMPAIGN'], size=(10,1),key='TOTAL_BOOKS_IN_CAMPAIGN'),
        sg.Push(),
        sg.Button("About", tooltip="Credits and info. TO BE IMPLEMENTED",disabled=True)
    ]
    row_final = [
        sg.Button('Save', key='-SAVE-PREFS-'),
        sg.Button('Cancel', key='-DONT-SAVE-PREFS-'),
        sg.Push(),
        sg.Button('Restore Defaults', key='-RESTORE-DEFAULT-PREFS-'),
    
    ]
    layout = [
        [row_1],
        [row_2],
        [row_3],
        [row_4],
        [row_5],
        [row_6],
        [row_final]
            ]
    
    return layout


def update_master_books_array(the_array):
    master_list_stats['TOTAL_UNIQUE_TITLES_IN_MASTER'] = the_array['rows']
    master_list_stats['TOTAL_VALUE_OF_SINGLE_UNIQUE_TITLES'] = the_array['market_value']
    master_list_stats['TOTAL_BOOKS_IN_MASTER'] = the_array['number_extant_copies']
    master_list_stats['TOTAL_BOOKS_IN_MASTER_FOR_PLACEMENT'] = the_array ['number_extant_available_to_place']

def zero_out_master_books_file():
    master_list_stats['TOTAL_UNIQUE_TITLES_IN_MASTER'] = 0
    master_list_stats['TOTAL_VALUE_OF_SINGLE_UNIQUE_TITLES'] = 0
    master_list_stats['TOTAL_BOOKS_IN_MASTER'] = 0
    master_list_stats['TOTAL_BOOKS_IN_MASTER_FOR_PLACEMENT'] = 0

    save_master_books_settings()

    sg.popup_notify("Master book settings have been zeroed out.",
                    title = "New master",
                    icon = radio_unchecked_icon,
                    display_duration_in_ms = config['duration_toaster_popups'],
                    fade_in_duration = config['fade_in_duration_toaster_popups'],
                    alpha = config['alpha_toaster_popups'],
                    location = None)

######################## CLASSES ########################

vocab_dictionary = import_language_words() # this is here because must come after definition of function
# read in dataframe for master file

master_book_pandas_table = read_excel_file_into_pandas (filename = "master_fantasy_book_list.xlsx", worksheet = "Master List")
master_excel_workbook, master_excel_worksheet = load_excel_objects(filename = "master_fantasy_book_list.xlsx", worksheet = "Master List")
stats = calculate_stats_excel(master_excel_workbook, master_excel_worksheet)

class FantasyBook():
    ''' 
    Fantasy book object.
    '''

    def __init__(self,
        book_type = 'Standard',         
        topic = '',
        topic_apparent = '',
        topic_title_form = '',
        book_title = '',
        book_title_flavor = '',
        cost_per_page = '',
        author_sex = '',
        author_name = '',
        author_title = '',
        author_epithet = '',
        author_full = '',
        author_nationality = '',
        current_language = '',
        original_language = '',
        is_a_translation = False,
        translator_name = '',
        translator_nationality = '',
        translator_sex = '',
        translator_title = '',
        translator_full_name = '',
        format = '',
        template = '',
        materials = '',
        libraries_it_is_in = '',
        number_extant_copies = 0,
        number_extant_available_to_place = 0,
        scope = 0,
        scope_esoteric = 0,
        complexity = 0,
        complexity_esoteric = 0,
        age_at_discovery = 0,
        number_pages = 0,
        reading_time = 0,
        reference_time = 0,
        production_value = 0,
        literary_value_base = 0,
        literary_value_modified = 0,
        esoteric_literary_value_base = 0,
        esoteric_literary_value_modified = 0,
        rarity_modifier = 0,
        weight = 0,
        number_volumes = 0,
        year_discovered = 0,
        year_written = 0,
        market_value = 0,
        weight_per_page = 0,
        fraction_complete = 0,
        uuid = '',
        note=''
        ):

        # set all values to whatever they were passed in 
        self.book_type = book_type
        self.topic = topic
        self.topic_apparent = topic_apparent
        self.topic_title_form = topic_title_form
        self.book_title = book_title
        self.book_title_flavor = book_title_flavor
        self.author_sex = author_sex
        self.author_name = author_name
        self.author_title = author_title
        self.author_epithet = author_epithet
        self.author_full = author_full
        self.author_nationality = author_nationality
        self.cost_per_page = cost_per_page
        self.current_language = current_language
        self.original_language = original_language
        self.is_a_translation = is_a_translation
        self.template = template
        self.translator_name = translator_name
        self.translator_nationality = translator_nationality
        self.translator_sex = translator_sex
        self.translator_title = translator_title
        self.translator_full_name = translator_full_name
        self.format = format
        self.materials = materials
        self.libraries_it_is_in = libraries_it_is_in
        self.number_extant_copies = number_extant_copies
        self.number_extant_available_to_place = number_extant_available_to_place
        self.scope = scope
        self.scope_esoteric = scope_esoteric
        self.complexity = complexity
        self.complexity_esoteric = complexity_esoteric
        self.age_at_discovery = age_at_discovery
        self.number_pages = number_pages
        self.reading_time = reading_time
        self.reference_time = reference_time
        self.production_value = production_value
        self.literary_value_base = literary_value_base
        self.literary_value_modified = literary_value_modified
        self.esoteric_literary_value_base = esoteric_literary_value_base
        self.esoteric_literary_value_modified = esoteric_literary_value_modified
        self.rarity_modifier = rarity_modifier
        self.weight = weight
        self.number_volumes = number_volumes
        self.year_discovered = year_discovered
        self.year_written = year_written
        self.market_value = market_value
        self.weight_per_page = weight_per_page
        self.fraction_complete = fraction_complete
        self.year_discovered = year_discovered
        self.year_written = year_written
        self.libraries_it_is_in = libraries_it_is_in
        self.uuid = uuid
        self.note = note

        ### functions

        self.scope_set(self.scope)
        self.current_language_set(self.current_language)
        self.age_set(self.age_at_discovery)
        
        # translator_set must be called before original_language_set
        self.translator_set(translator_name=self.translator_name,translator_sex = self.translator_sex,translator_nationality=self.translator_nationality, translator_title=self.translator_title,translator_full_name=self.translator_full_name) 
        
        self.original_language_set(original_language = self.original_language, is_a_translation= self.is_a_translation)
        self.topic_set(topic = self.topic)
        self.topic_title_set(topic_title_form = self.topic_title_form)
        self.sex_set(author_sex = self.author_sex)
        self.author_epithet_set (author_epithet = self.author_epithet)
        self.author_name_set(author_name = self.author_name)
        self.author_title_set (author_title = self.author_title)
        self.author_full_set (author_full = self.author_full)
        self.complexity_set(complexity = self.complexity)
        self.format_set(format = self.format)
        self.book_title_set(book_title = self.book_title)
        self.materials_set(materials = self.materials)
        self.rarity_set(rarity_modifier = self.rarity_modifier, number_extant_copies = self.number_extant_copies, number_extant_available_to_place = self.number_extant_available_to_place)

        # these are all calculated based on other values above
        self.number_pages_set(number_pages = self.number_pages)
        self.reading_time_set(reading_time = self.reading_time)
        self.production_value_set(production_value = self.production_value)
        self.literary_value_set()
        self.weight_set(weight = self.weight)
        self.volumes_number_set(number_volumes = self.number_volumes)
        self.flavor_text_title_set(book_title_flavor = self.book_title_flavor)
        self.percentage_of_text_missing_set(fraction_complete = self.fraction_complete)
        self.esoteric_set()
        self.uuid_create()

    def add(self,library):
        ''' Add this book to a given library'''
        self.libraries_it_is_in.append(library)

    def age_set(self,age=None):
        if not age:
            table_name = 'BookAge_'+ self.current_language # Ancient, Dwarvish, Elvish, Classical, Common are options
            dice_string = self.book_details_result_from_tables(table_name)
            if self.is_a_translation == True: 
                self.age_at_discovery = d20.roll(preferences['TRANSLATION_ADDITIONAL_AGE_OF_ORIGINAL']).total # bonus to age if is translation.
            
            self.age_at_discovery = self.age_at_discovery + d20.roll(dice_string).total
            if self.age_at_discovery < preferences['MINIMUM_AGE_BOOK']: self.age_at_discovery = preferences['MINIMUM_AGE_BOOK']
            if self.age_at_discovery > preferences['MAXIMUM_AGE_BOOK']: self.age_at_discovery = preferences['MAXIMUM_AGE_BOOK']
        else:
            self.age_at_discovery = age
            self.age = age
            if self.age < preferences['MINIMUM_AGE_BOOK']: self.age = preferences['MINIMUM_AGE_BOOK']
            if self.age > preferences['MAXIMUM_AGE_BOOK']: self.age = preferences['MAXIMUM_AGE_BOOK']
    
    def add_note (self,note=None):
        self.note = note

    def append_note(self,note=None):
            '''
            Adds the note text to the existing text already present.
            '''

            if note:
                if self.note != '': self.note = self.note + " " # add a space if any text in the field
                self.note = self.note + str(note)

    def author_epithet_set (self, author_epithet=None):
        if not author_epithet:
            if preferences['CHANCE_OF_EPITHET_IN_AUTHOR_NAME'] > d20.roll("1d100").total:
                author_epithet = nt['epithets_table'].df.sample() # a random option is then chosen
                author_epithet = author_epithet.iloc[0,0]         
        self.author_epithet = author_epithet            

    def author_full_set (self, author_full=None):
        # put it all together
        if not author_full:

            if self.author_title != "" and self.author_title != "None": author_full = author_full.join([self.author_title," "])
            author_full += (self.author_name)
            if self.author_epithet != "" and self.author_epithet != "None": author_full = author_full + " " + self.author_epithet
        
        self.author_full = author_full
    
    def author_name_set(self,author_name=None,author_nationality = None):
        
        if not author_name:        
             
            author_name, author_nationality, _ = self.name_generate(sex = self.author_sex)
            
        self.author_name = author_name
        self.author_nationality = author_nationality

    def author_title_set(self, author_title=None):
        if not author_title:
            self.author_title = self.person_title_generate(sex = self.author_sex)
        else:
            self.author_title = author_title

    def book_details_result_from_tables(self,table_for_value,roll_result=None):
        ''' Checks table aaDiceTypeToRoll to see what dice to roll, and then rolls and checks the result on the given table.
        All tables that use this should have:
        1) An entry in SQL table aaDiceTypeToRoll with the same table name that matches table to roll on.
        2) The SQL table to be rolled on should have only two columns: one called 'DieRange' and the other 'Result'. 

        If roll_result is passed in as an integer value, then the table looks up as if that roll had been made (i.e., not random).
        '''

        # what dice to roll on the table   
       
        if not roll_result:
            dice_string = r.what_dice_to_roll(table_for_value) # returns a list
            
            if dice_string == []:
                raise ValueError("Empty dice list returned -- does SQL table 'aaDiceTypeToRoll' have an entry for this table?")
            else:
                roll_result = r.d20.roll(dice_string).total 

        # actually pass the roll value now that we know what dice have given
        t = r.RPG_table(table_for_value)
        rolled_row = t.roll(roll_result)
        return rolled_row 
    
    def book_title_set(
            self,
            adjective_1=None, 
            noun_1=None, 
            noun_2=None, 
            study_of=None, 
            study_in=None, 
            study_on=None, 
            template=None,
            book_title=None,
            conjunction_about=None,
            conjunction_by=None,
            negative_1=None,
            place_city=None,
            place_nation=None,
            religious_starter=None,
            verbing=None,
            saint=None,
            person_1=None,
            person_2=None,
            person_famous=None,
            communication=None,
            biography_starter=None,
            person_evil=None,
            history_of=None,
            the_1=None,

            ):

        avoid_special_class_of_title = True

        if book_title: 
            self.book_title = book_title
            self.template = "Final title passed in as: " + book_title
        
        else:
            if not template: template = nt['titles_template_list_general'].df.sample().iloc[0,0]
            topic = self.topic_title_form

            if "theology" in topic.lower():
                while ("religious" not in template) and ("biography" not in template):
                    template = nt['titles_template_list_theology'].df.sample().iloc[0,0]

            if "history" in topic.lower():
                while ("history" not in template) and ("biography" not in template):
                    template = nt['titles_template_list_history'].df.sample().iloc[0,0]

            if ("occult" in topic.lower()) or ("apostasy" in topic.lower()) or ("black lore" in topic.lower()):
                while ("occult" not in template) and ("negative" not in template) and {"evil" not in template} and ("biography" not in template):
                    template = nt['titles_template_list_occult'].df.sample().iloc[0,0]
            
            # refactor this code eventually into loop with eval()

            if not adjective_1 and "{adjective_1}" in template: adjective_1 = nt['titles_adjective_1_list'].df.sample().iloc[0,0]
            if not noun_1 and "{noun_1}" in template: noun_1 = nt['titles_noun_1_list'].df.sample().iloc[0,0]
            if not noun_2 and "{noun_2}" in template: noun_2 = nt['titles_noun_2_list'].df.sample().iloc[0,0]
            if not study_of and "{study_of}" in template: study_of = nt['titles_study_of_list'].df.sample().iloc[0,0]
            if not study_in and "{study_in}" in template: study_in = nt['titles_study_in_list'].df.sample().iloc[0,0]
            if not study_on and "{study_on}" in template: study_on = nt['titles_study_on_list'].df.sample().iloc[0,0]
            if not conjunction_about and "{conjunction_about}" in template: conjunction_about = nt['titles_conjunction_about'].df.sample().iloc[0,0]
            if not conjunction_by and "{conjunction_by}" in template: conjunction_by = nt['titles_conjunction_by'].df.sample().iloc[0,0]
            if not negative_1 and "{negative_1}" in template: negative_1 = nt['titles_negative_subject'].df.sample().iloc[0,0]
            if not place_city and "{place_city}" in template: place_city = nt['titles_places_cities'].df.sample().iloc[0,0]
            if not place_nation and "{place_nation}" in template: place_nation = nt['titles_places_nations'].df.sample().iloc[0,0]
            if not religious_starter and "{religious_starter}" in template: religious_starter = nt['titles_religious_starter'].df.sample().iloc[0,0]
            if not verbing and "{verbing}" in template: verbing = nt['titles_study_verbing'].df.sample().iloc[0,0]
            if not saint and "{saint}" in template: saint = nt['titles_saints_amalgamated'].df.sample().iloc[0,0]
            if not person_famous and "{person_famous}" in template: person_famous = nt['titles_person_famous_amalgamated'].df.sample().iloc[0,0]
            if not communication and "{communication}" in template: communication = nt['titles_communication'].df.sample().iloc[0,0]
            if not biography_starter and "{biography_starter}" in template: biography_starter = nt['titles_biography_starter'].df.sample().iloc[0,0]
            if not person_evil and "{person_evil}" in template: person_evil = nt['titles_person_evil'].df.sample().iloc[0,0]
            if not person_1 and "{person_1}" in template: person_1, _ , _ = self.name_generate() # 2nd,3rd are nation, sex which we don't need
            if not person_2 and "{person_2}" in template: person_2, _, _ = self.name_generate()
            if not history_of and "{history_of}" in template: history_of = nt['titles_history_of'].df.sample().iloc[0,0]
            if not the_1 and "{the_1}" in template: the_1 = nt['titles_the_1'].df.sample().iloc[0,0]
            
            self.template = template

            self.book_title = template.format(
                adjective_1 = adjective_1, 
                noun_1 = noun_1, 
                noun_2 = noun_2, 
                study_of = study_of, 
                study_in = study_in, 
                study_on = study_on, 
				topic = topic,
				conjunction_about=conjunction_about,
				conjunction_by=conjunction_by,
				negative_1=negative_1,
				place_city=place_city,
				place_nation=place_nation,
				religious_starter=religious_starter,
				verbing=verbing,
				saint=saint,
				person_1=person_1,
				person_2=person_2,
				person_famous=person_famous,
				communication=communication,
				biography_starter=biography_starter,
				person_evil=person_evil,
				history_of=history_of,
                the_1 = the_1
            )

    def complexity_set(self,complexity=None):
        if not complexity:
            complexity_from_table = self.book_details_result_from_tables(config['complexity_table_list'][self.scope-1]) # Minus 1 since list index starts at zero.
            if complexity_from_table >= 1: complexity_from_table = int(complexity_from_table) # doesn't integerize 0.75
            if complexity_from_table - self.scope > 4:
                complexity_from_table = 5-self.scope

            self.complexity = complexity_from_table
        
        else:
            self.complexity = complexity   
    
    def current_language_set(self, current_language = None):
        if not current_language:
            self.current_language = self.book_details_result_from_tables('BookCurrentLanguage')
        else:
            self.current_language = current_language

    def delete_note (self):
        self.add_note(note='')

    def esoteric_set(self):
        if self.topic != self.topic_apparent:
            esoteric_complexity_from_table = 0
            esoteric_ratios_correct = False
            count_times_through = 0 # don't let loop through too much until we take matter into our hands and define it. :-)

            while not esoteric_ratios_correct:
                # esoteric scope
                self.scope_esoteric = self.book_details_result_from_tables('BookScope')
                
                # esoteric complexity
                esoteric_complexity_from_table = self.book_details_result_from_tables(config['complexity_table_list'][int(self.scope_esoteric)-1])
                if esoteric_complexity_from_table >= 1: esoteric_complexity_from_table = int(esoteric_complexity_from_table)      
                self.complexity_esoteric = esoteric_complexity_from_table

                # Is apparent ratio >= to the esoteric?
                if self.scope < 1: self.scope = 1 # self.scope can be very low if lots of the book is missing. Corrected below

                ratio_apparent = self.scope/self.complexity
                ratio_esoteric = self.scope_esoteric/self.complexity_esoteric

                if ratio_apparent >= ratio_esoteric: 
                    esoteric_ratios_correct = True
                    self.scope = self.scope * self.fraction_complete # restores from previous set = 1.
                    
                else:
                    if count_times_through >= 10:
                        self.scope = self.scope_esoteric
                        self.complexity = self.complexity_esoteric
                        self.scope = self.scope * self.fraction_complete
                        esoteric_ratios_correct = True
    
            self.esoteric_value_set()

    def esoteric_value_set (self):
        target_table = 'BookLiteraryValueScope' + str(int(self.scope_esoteric))
        self.complexity_esoteric = self.hack_complexity(target_table, complexity=self.complexity_esoteric)
        self.esoteric_literary_value_base = self.look_up_table(
            table_name=target_table,
            search_column='Complexity',
            search_term = self.complexity_esoteric,
            result_column='LiteraryValue'
            )
        self.esoteric_literary_value_modified = ceil (self.esoteric_literary_value_base * self.rarity_modifier) * self.number_pages * 5 # Writer of 18 Intelligence needed; hence the 5.
        self.market_value = ceil(self.literary_value_modified + self.esoteric_literary_value_modified + self.production_value)

    def format_set(self, format = None):
        if not format:
            target_table = 'BookAge_Format_'
            if self.age_at_discovery < 11: target_table += '0001_0010'
            elif self.age_at_discovery < 51: target_table += '0011_0050'
            elif self.age_at_discovery < 101: target_table += '0051_0100'
            elif self.age_at_discovery < 501: target_table += '0101_0500'
            elif self.age_at_discovery < 1001: target_table += '0501_1000'
            elif self.age_at_discovery < 2001: target_table += '1001_2000'
            elif self.age_at_discovery < 10001: target_table += '2001_10000'

            self.format = self.book_details_result_from_tables(target_table)
        else:
            self.format = format
    
    def flavor_text_title_set(self, book_title_flavor=None):
        
        if not book_title_flavor:

            # Limit number chars (like Akkadian, gothic_latin)
            if  self.current_language in config['lang_limit_40_chars']:
                limit_chars = 40 # These require only 40 chars or weird stuff happens. ? Unicode issue
            else:
                limit_chars = 0 # all the rest no limit

            # no spaces between:
            if self.current_language in config['lang_no_spaces']:
                spaces = False # No spaces between works; Kanji looks better, for example.
            else:
                spaces = True # all the rest have spaces

            if self.current_language == "Common":
                book_title_flavor = self.book_title
            else:
                try:
                    
                    num_words_in_english_title = len(self.book_title.split())
                    num_words_in_flavor_title = eval(config['DEFAULT_FORMULA_CALC_NUM_FLAV_TEXT_WORDS_FROM_ORIG_TITLE'])
                    if num_words_in_flavor_title <3: num_words_in_flavor_title = eval(config['DEFAULT_FLAVOR_TEXT_NUMBER_OF_WORDS'])

                    book_title_flavor = str(
                        lf.words(vocab_dictionary[self.current_language],
                        count=num_words_in_flavor_title,
                        limit=limit_chars,
                        spaces = spaces)
                    )

                except:
                    book_title_flavor = "No flavor text designated for this language type."
                book_title_flavor = book_title_flavor.capitalize()

        self.book_title_flavor = book_title_flavor
    
    def hack_complexity(self, target_table,complexity):
        ''' Prevents occasional bizarre complexity value causing lookup errors in the SQL table. This is a disgraceful, hacky solution pending  a better understanding of where the occasional bug is coming from. '''

        if complexity > 4 and target_table == "BookLiteraryValueScope1":
            return 4
        elif complexity > 5 and target_table == "BookLiteraryValueScope2":
            return 5
        elif complexity > 6 and target_table == "BookLiteraryValueScope3":
            return 6
        elif complexity > 7 and target_table == "BookLiteraryValueScope4":
            return 7
        else: 
            return complexity
        
    def literary_value_set (self):
        
        if self.scope >= 1:
            target_table = 'BookLiteraryValueScope' + str(int(self.scope))
        else:
            target_table = 'BookLiteraryValueScope' + str('1')
        self.complexity = self.hack_complexity(target_table, complexity=self.complexity)

        self.literary_value_base = self.look_up_table(
            table_name=target_table,
            search_column='Complexity',
            search_term = self.complexity,
            result_column='LiteraryValue'
            )

        self.literary_value_modified = ceil(self.literary_value_base * self.rarity_modifier * self.number_pages)
        self.market_value = ceil(self.literary_value_modified + self.production_value)

    def look_up_table (self,result_column,table_name,search_column,search_term):
        query = 'SELECT {} from {} where {} LIKE "{}"'.format(result_column,table_name,search_column,search_term)
        t = r.LookUpTable(query = query)
        return t.result   
    
    def materials_set (self, materials=None):
        
        if not materials:
            target_table = 'BookMaterials' + self.format
            self.materials = self.book_details_result_from_tables(target_table)

        else:
            self.materials = materials
   
    def name_generate(self,sex=None):
        # first name
        if sex == None:
            if d20.roll("1d100").total <  preferences['CHANCE_OF_FEMALE_AUTHOR']: 
                sex = "Female"
            else: 
                sex = "Male"

        if sex == "Male": first_name = nt['complete_table_male_names'].df.sample()
        else: first_name = nt['complete_table_female_names'].df.sample()
        author_nationality = (first_name.iloc[0,1]) # the second column (i.e. index 1 since starts at 0) is the table for this type of name's surname.
                
        # surname
        last_name_table = (nt['surnames_tables'][author_nationality])
        last_name = last_name_table.df.sample()
        author_name = str(first_name.iloc[0,0]) + " " + str(last_name.iloc[0,0]) # first (0 index) item is the name

        return author_name, author_nationality, sex
    
    def number_pages_set(self, number_pages = None):
        if number_pages:
            self.number_pages = number_pages
            self.complexity = ceil((self.scope * 1000) / self.number_pages)

        else:
            self.number_pages = ceil((self.scope * 1000) / self.complexity) # note integer division // 

    def original_language_set(self, original_language=None,is_a_translation=False):
        
        if is_a_translation:
            self.is_a_translation = is_a_translation

        if (not is_a_translation) and (not original_language): 
            self.is_a_translation == False
            self.original_language == ''
            return

        if not original_language: # original language is empty
            original_language = self.book_details_result_from_tables('BookOriginalLanguage')

            while original_language == self.current_language:  
                original_language = self.book_details_result_from_tables('BookOriginalLanguage') # reroll until not the same
        
        self.original_language = original_language
        self.is_a_translation = True

    def percentage_of_text_missing_set(self,fraction_complete=None):
        
        if not fraction_complete:
            if preferences['CHANCE_OF_INCOMPLETE_WORK'] >= d20.roll("1d100").total:
                fraction_missing = round(d20.roll("1d99").total/100,2)
            else:
                fraction_missing = 0
            fraction_complete = 1 - fraction_missing
         
        self.scope = ceil(self.scope * 2.0 * fraction_complete) / 2.0 # the x2, then div 2 rounds to nearest 0.5
        if self.scope < 1: self.scope = 1

        self.reading_time = round(self.reading_time * 2.0 * fraction_complete) / 2.0
        self.reference_time = round(self.reference_time * 2.0 * fraction_complete) / 2.0

        self.number_pages = ceil(self.number_pages * fraction_complete)
        self.weight = ceil(self.weight * fraction_complete)
        self.market_value = ceil(self.market_value * fraction_complete)

        self.fraction_complete = round(fraction_complete,2)
    
    def person_title_generate (self,sex="Male"):
        
        author_title = ''

        if preferences['CHANCE_OF_TITLE_IN_AUTHOR_NAME'] >= d20.roll("1d100").total:

            author_title = str(nt['author_title_table'].df.sample().iloc[0,0])
           
        #  # male/female titles are separated by a slash in the SQL database  
            if author_title.__contains__("/"):
                title_split = author_title.split("/",2)
                
                if sex == "Male":
                    author_title = title_split[0]
                else:
                    author_title = title_split[1]
        
        return string.capwords(str(author_title))
    
    def production_value_set(self, production_value = None):

        if production_value:
            self.production_value = production_value
            self.cost_per_page = production_value / self.number_pages

        else:
            target_table = "BookProductionValue" + self.format
            self.cost_per_page = self.look_up_table(result_column="Cost",table_name=target_table,search_column="Material",search_term=self.materials)
            self.production_value = ceil(self.cost_per_page * self.number_pages)
        
    def rarity_set(self, rarity_modifier = None, number_extant_copies = None, number_extant_available_to_place = None):
        the_roll = d20.roll("1d100").total # this same value needed twice, so must roll it first so can be passed.

        if number_extant_copies:
            self.number_extant_copies = number_extant_copies
        
        else:
            dice_string_determine_number_copies = self.book_details_result_from_tables("BookRarityCopies", roll_result= the_roll)
            number_of_copies_roll = d20.roll(dice_string_determine_number_copies).total
            self.number_extant_copies = number_of_copies_roll
        
        if number_extant_available_to_place:
            self.number_extant_available_to_place = number_extant_available_to_place

        else:
            self.number_extant_available_to_place = (self.number_extant_copies - 1) # ie, less this one.

        if rarity_modifier:
            self.rarity_modifier = rarity_modifier
        
        else:
            self.rarity_modifier = self.book_details_result_from_tables("BookRarityModifier",roll_result=the_roll)

    def reading_time_set(self,reading_time = None):

        if reading_time:
            self.number_pages = config['READING_PAGES_PER_HOUR'] * reading_time
            self.reading_time = reading_time
            self.reference_time = reading_time

        else:
            self.reading_time = ceil(self.number_pages/config['READING_PAGES_PER_HOUR'])
            self.reference_time = self.reading_time

    def refresh_number_pages(self):
            self.number_pages_set(number_pages = self.number_pages)
            self.reading_time_set()
            self.production_value_set()    

    def remove (self,library):
        ''' Remove this book from a given library'''
        try:
            self.libraries_it_is_in.remove(library)
        except ValueError:
            print ("The book _{}_ is not in {} library.".format(self.title, library))
    
    def scope_set(self, scope=None):
        if not scope:
            self.scope = self.book_details_result_from_tables("BookScope")
        else:
            self.scope = scope
    
    def sex_set (self, author_sex=None):
        if not author_sex:
            if d20.roll("1d100").total <= preferences['CHANCE_OF_FEMALE_AUTHOR']: self.author_sex = "Female"
            else: self.author_sex = "Male"
        else:
            self.author_sex = author_sex
    
    def topic_set (self, topic=None):

        if not topic:
            topic = self.book_details_result_from_tables("BookTopicsACKS")
           
        self.topic = topic
        self.topic_apparent = topic     
        
        if "Esoteric" in self.topic:
                while "Esoteric" in topic: # keep picking apparent topic until not esoteric
                    topic = self.book_details_result_from_tables("BookTopicsACKS")
                    
                self.topic_apparent = topic
                  
    def topic_title_set(self,topic_title_form=None):
        if not topic_title_form:

            t = self.look_up_table(
                result_column="title_string",
                table_name= "_book_titles_topics", 
                search_column="Result",
                search_term = self.topic_apparent
                )
            
            t = t.split(";") # list is made by separating by semicolons
            t = random.choice(t) # a random option is then chosen
            self.topic_title_form = t
        
        else:
            self.topic_title_form = topic_title_form

    def translator_set (self,translator_name=None,translator_sex = None,translator_nationality=None, translator_title=None,translator_full_name=None):
        roll_to_see_if_it_is_a_translation = d20.roll("1d100").total

        self.is_a_translation = False
        self.translator_full_name = translator_full_name

        if (translator_name or translator_full_name) and (self.current_language not in preferences['ANCIENT_LANGUAGES_WHICH_WOULD_NOT_BE_TRANSLATED_INTO']):
            self.is_a_translation = True
            self.translator_nationality = translator_nationality
            self.translator_title = translator_title
            self.translator_name = translator_name
            self.translator_sex = translator_sex

            if translator_full_name:
                self.translator_full_name = translator_full_name
            else:
                self.translator_full_name = self.translator_title + " " + self.translator_name

        elif (roll_to_see_if_it_is_a_translation < preferences['CHANCE_OF_BEING_TRANSLATION']) and (self.current_language not in preferences['ANCIENT_LANGUAGES_WHICH_WOULD_NOT_BE_TRANSLATED_INTO']):
            self.is_a_translation = True
            self.translator_name, self.translator_nationality, self.translator_sex = self.name_generate()
            self.translator_title = self.person_title_generate(sex = self.translator_sex)
            self.translator_full_name = self.translator_title + " " + self.translator_name

        else:
            
            self.is_a_translation = False

    def uuid_create (self):
        self.uuid = str(uuid.uuid4())

    def volumes_number_set(self, number_volumes = None):
        if number_volumes:
            self.number_volumes = number_volumes

            if self.format == "Codex":
                self.number_pages = ceil (self.number_volumes * config['PAGES_PER_VOLUME_FOR_CODEX'])
                self.weight = ceil ((self.number_pages * self.weight_per_page) + (self.number_volumes * config['WEIGHT_PER_VOLUME_OF_CODEX']))
            
            elif self.format == "Scroll":
                self.number_pages = ceil (self.number_volumes * config['PAGES_PER_VOLUME_FOR_SCROLL'])
                self.weight = ceil((self.number_pages * self.weight_per_page) + (self.number_volumes * config['WEIGHT_PER_VOLUME_OF_SCROLL']))

            elif self.format == "Tablet":
                self.number_volumes = 1 # ie, never multivolume
            
            else:
                raise ValueError("Format has a problem: is not a Codex, Scroll, or Tablet.")

            self.refresh_number_pages()

        else:

            if self.format == "Codex":
                self.number_volumes = ceil(self.number_pages/config['PAGES_PER_VOLUME_FOR_CODEX'])
                self.weight = self.weight + (self.number_volumes * config['WEIGHT_PER_VOLUME_OF_CODEX'])
            
            elif self.format == "Scroll":
                self.number_volumes = ceil(self.number_pages/config['PAGES_PER_VOLUME_FOR_SCROLL'])
                self.weight = self.weight + (self.number_volumes * config['WEIGHT_PER_VOLUME_OF_SCROLL'])
            
            elif self.format == "Tablet":
                self.number_volumes = 1 # ie, never multivolume

            else:
                raise ValueError("Format has a problem: is not a Codex, Scroll, or Tablet.")
        
    def weight_set(self,weight=None):
        self.weight_per_page = self.look_up_table(result_column="Result",table_name="BookWeight",search_column="Material",search_term=self.materials)

        if weight:
            self.weight = weight
            self.number_pages  = ceil(self.weight / self.weight_per_page)
            self.refresh_number_pages()

        else:
            self.weight = ceil(self.weight_per_page * self.number_pages)

class MagicBook(FantasyBook):
    ''' Subclass of fantasy book, that has a few extra values.'''
    def __init__ (self,
        book_type = "Magic",
    ):
        super().__init__(self)


######################## main() ########################

sg.theme('Dark Blue 3')
window1 = sg.Window(
    'Fantasy Books Generator', 
    layout = fantasy_books_main_gui(),
    grab_anywhere = True,
    resizable = False,
    icon = books_icon,
    finalize = True
    )

sg.theme("Dark Blue 12")
window_settings = sg.Window(
    'Preferences',
    layout = settings_gui(),
    grab_anywhere=True,
    icon = settings_general_icon,
    finalize=True,
    disable_close = True,
    modal = False,
)

# TO BE IMPLEMENTED 

# sg.theme("Dark Green 1")
# window_about = sg.Window(
#     'Preferences',
#     layout = about_window_gui(),
#     grab_anywhere=True,
#     icon = settings_general_icon,
#     finalize=True,
#     disable_close = False,
#     modal = False,
# )

# window2 = sg.Window(
#     'Fantasy Books Generator', 
#     layout = progress_window_gui(),
#     grab_anywhere = True,
#     alpha_channel = 0.7,
#     no_titlebar=True,
#     resizable = False,
#     # icon = books_icon,
#     finalize = True
#     )

window_settings.hide()
# window2.move(window1.current_location()[0]+500, window1.current_location()[1]+200)

# turn off tabbing to all elements in window1
for element in window1.key_dict.values():
        element.block_focus()

# retore tabbing to some in window1
window1['-number_of_books_to_make-'].block_focus(block=False)
window1['-value_of_books_to_make-'].block_focus(block=False)
########## Main Event Loop of GUI

while True:
    window,event, values = sg.read_all_windows()
    if event in (sg.WIN_CLOSED, 'Quit'):
        break
    
    elif event == "-value_of_books_to_make-": # only allows integers, does not allow to be blank
         if len(values['-value_of_books_to_make-']) > 0:
            if values['-value_of_books_to_make-'][-1] not in ('0123456789'):
                window1['-value_of_books_to_make-'].update(values['-value_of_books_to_make-'][:-1])
         else:
             window1['-value_of_books_to_make-'].update('0')
    
    elif event == "-number_of_books_to_make-": # only allows integers, does not allow to be blank
         if len(values['-number_of_books_to_make-']) > 0:
            if values['-number_of_books_to_make-'][-1] not in ('0123456789'):
                window1['-number_of_books_to_make-'].update(values['-number_of_books_to_make-'][:-1])
         else:
             window1['-number_of_books_to_make-'].update('0')

    elif event == 'Overshoot':                # if the normal button that changes color and text
            
            overshoot_toggle = overshoot_event(overshoot_toggle = overshoot_toggle)

    elif event == 'Edit preferences':
        window1.hide()
        window_settings.un_hide()

    elif event == 'Save settings and Quit':
        save_gui_settings()
        break
    
    elif event == 'Clear master Excel file':
        zero_out_master_books_file()
        backup_excel_file(filename = 'master_fantasy_book_list.xlsx')
        create_new_master_excel_file(filename = 'master_fantasy_book_list.xlsx')

    elif event == 'Generate Books':
        print ("Start:" + str(time.asctime()))
        window1.set_cursor("watch")
        # window1.hide()
        save_gui_settings()


        excel_filename = values['-EXCEL_OUT_FILENAME-']
        excel_worksheet = values['-EXCEL_OUT_WORKSHEET-']
        master_filename = values['-MASTER_FILENAME-']
        master_worksheet = values['-MASTER_WORKSHEET-']
        value_of_books = int(values['-value_of_books_to_make-'])
        number_of_books = int(values['-number_of_books_to_make-'])

        if window1['-R1-'].metadata:
            books, books_value = book_hoard (
                value_of_books=value_of_books,
                overshoot=overshoot_toggle, 
                )
            
        else:
            books, books_value = book_batch(
                number = number_of_books,
                )
        
        export_books_to_excel(
            books,
            filename = excel_filename, 
            worksheet = excel_worksheet)

        master_excel_workbook.save('master_fantasy_book_list.xlsx') # save the master list with the decremented number of books for that title.
        master_excel_workbook.close()

        archive_to_master(
            source=excel_filename, 
            source_worksheet = excel_worksheet,
            destination=master_filename,
            destination_worksheet = master_worksheet,
            )

        master_as_pandas = read_excel_file_into_pandas(
            filename = master_filename,
            worksheet = master_worksheet,
            )
        # TO_DO
        master_excel_workbook, master_excel_worksheet = load_excel_objects(filename = 'master_fantasy_book_list.xlsx', worksheet = 'Master List')
        stats = calculate_stats_excel(master_excel_workbook, master_excel_worksheet)
        update_master_books_array(stats)
        save_master_books_settings() # save data for next time.

        # load for next round:
        master_book_pandas_table = read_excel_file_into_pandas(
            filename = 'master_fantasy_book_list.xlsx',
            worksheet = 'Master List',
            )
        window1.set_cursor("arrow")
        window1.un_hide()
        print ("End:" + str(time.asctime()))

    elif event == "Reset to defaults":
        window1['-EXCEL_OUT_FILENAME-'].update(value="books_spreadsheet_out.xlsx")
        window1['-EXCEL_OUT_WORKSHEET-'].update(value="Book Hoard")
        window1['-MASTER_FILENAME-'].update(value="master_fantasy_book_list.xlsx")
        window1['-MASTER_WORKSHEET-'].update(value="Master List")
        window1['-value_of_books_to_make-'].update(value = 0)
        window1['-number_of_books_to_make-'].update(value = 0)
        window1['-R1-'].update(radio_checked_icon)
        window1['-R1-'].metadata = True
        window1['-R2-'].update(radio_unchecked_icon)
        window1['-R2-'].metadata = False

    elif event in radio_keys:
            check_radio(event)
    
    elif event.startswith('-T'):        # If text element clicked, change it into a radio button key
        check_radio(event.replace('T', 'R'))

    elif event == 'Clear_History_Default_Out':
        sg.user_settings_set_entry('-default_out_filenames-', [])
        sg.user_settings_set_entry('-last_default_out_filename-', '')
        window1['-EXCEL_OUT_FILENAME-'].update(values=[], value='')

    elif event == 'Clear_Master_History':
        sg.user_settings_set_entry('-default_master_filenames-', [])
        sg.user_settings_set_entry('-last_default_master_filename-', '')
        window1['-MASTER_FILENAME-'].update(values=[], value='')

    elif event == 'Clear_History_Default_Out_Worksheet':
        sg.user_settings_set_entry('-default_out_worksheets-', [])
        sg.user_settings_set_entry('-last_default_out_worksheet-', '')
        window1['-EXCEL_OUT_WORKSHEET-'].update(values=[], value='')

    elif event == 'Clear_History_Master_Worksheet':
        sg.user_settings_set_entry('-default_master_worksheets-', [])
        sg.user_settings_set_entry('-last_default_master_worksheet-', '')
        window1['-MASTER_WORKSHEET-'].update(values=[], value='')
    
    elif event == "-SAVE-PREFS-":
        for the_setting in config['prefs_list_integers']:
            preferences[the_setting] = int(values[the_setting])

        for the_setting in config['prefs_list_strings']:
            preferences[the_setting] = values[the_setting]

        with open("preferences_fantasy_books.yaml", "w") as f:     
                    yaml.dump(preferences, stream=f, default_flow_style=False, sort_keys=True)

        window_settings.hide()
        window1.un_hide()
        
        sg.popup_notify("Preferences have been saved.",
                    title = "Settings saved!",
                    icon = settings_save_icon,
                    display_duration_in_ms = config['duration_toaster_popups_longer'],
                    fade_in_duration = config['fade_in_duration_toaster_popups'],
                    alpha = 0.9,
                    location = None)

    elif event == "-RESTORE-DEFAULT-PREFS-":
        for index,value in enumerate(config['prefs_list_integers']):
            window[value].update((config['prefs_list_integers_defaults'])[index])

        for index,value in enumerate (config['prefs_list_strings']):
               window[value].update((config['prefs_list_strings_defaults'])[index])
        
        sg.popup_notify("Default settings restored. You may Quit to keep, or Quit to discard and revert.",
                    title = "Default restored",
                    icon = settings_reset_icon,
                    display_duration_in_ms = config['duration_toaster_popups_longer'],
                    fade_in_duration = config['fade_in_duration_toaster_popups'],
                    alpha = 0.9,
                    location = None)

    elif event == '-DONT-SAVE-PREFS-':
        for index,value in enumerate(config['prefs_list_integers']):
            window[value].update(preferences[value])

        for index,value in enumerate (config['prefs_list_strings']):
               window[value].update(preferences[value])
        
        window_settings.hide()
        window1.un_hide()
        
        sg.popup_notify("Preferences were NOT saved. No changes were made.",
                    title = "Settings not saved.",
                    icon = settings_cancel_icon,
                    display_duration_in_ms = config['duration_toaster_popups_longer'],
                    fade_in_duration = config['fade_in_duration_toaster_popups'],
                    alpha = 0.9,
                    location = None)

window.close()

