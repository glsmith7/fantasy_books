import d20
from lorem_text_fantasy import lorem as lf
from math import ceil
from openpyxl.styles import Font as openpyxl_font
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import random as random
import rpg_tables as r
import string as string
import sys

# logging boilerplate
import settings_GLS as s
import logging
import logging_tools_GLS
logger = logging.getLogger(__name__)

################ GLOBALS #####################
global CHANCE_OF_BEING_TRANSLATION, TRANSLATION_ADDITIONAL_AGE_OF_ORIGINAL, ANCIENT_LANGUAGES_WHICH_WOULD_NOT_BE_TRANSLATED_INTO 
global CHANCE_OF_EPITHET_IN_AUTHOR_NAME, CHANCE_OF_TITLE_IN_AUTHOR_NAME, CHANCE_OF_FEMALE_AUTHOR
global WEIGHT_PER_VOLUME_OF_CODEX, WEIGHT_PER_VOLUME_OF_SCROLL
global READING_PAGES_PER_HOUR, PAGES_PER_VOLUME_FOR_CODEX, PAGES_PER_VOLUME_FOR_SCROLL
global CHANCE_OF_INCOMPLETE_WORK
global DEFAULT_FLAVOR_TEXT_NUMBER_OF_WORDS, DEFAULT_FORMULA_CALC_NUM_FLAV_TEXT_WORDS_FROM_ORIG_TITLE
global DEFAULT_EXCEL_FONT, DEFAULT_EXCEL_FLAVOR_FONT_SIZE
global vocab_dictionary
global lang_no_spaces, lang_limit_40_chars


vocab_dictionary = {}

#########################################################
# USER SETABLE variables
#########################################################
# ACKS given values
READING_PAGES_PER_HOUR = 180
PAGES_PER_VOLUME_FOR_CODEX = 750
PAGES_PER_VOLUME_FOR_SCROLL = 250
WEIGHT_PER_VOLUME_OF_CODEX = 1.5 # lbs
WEIGHT_PER_VOLUME_OF_SCROLL = 2 # lbs

#
ANCIENT_LANGUAGES_WHICH_WOULD_NOT_BE_TRANSLATED_INTO = ['Ancient']
CHANCE_OF_BEING_TRANSLATION = 5 # # 0-100%
CHANCE_OF_EPITHET_IN_AUTHOR_NAME = 15 # 0-100%
CHANCE_OF_TITLE_IN_AUTHOR_NAME = 30 # 0-100%
CHANCE_OF_FEMALE_AUTHOR = 50 # 0-100%
TRANSLATION_ADDITIONAL_AGE_OF_ORIGINAL = '1d100+20'
CHANCE_OF_INCOMPLETE_WORK = 5 # 0-100%

# uses first:
DEFAULT_FORMULA_CALC_NUM_FLAV_TEXT_WORDS_FROM_ORIG_TITLE='num_words_in_english_title - d20.roll("1d4").total + d20.roll("1d8").total'

# if the above gives less than 3 words, this formula is used instead
DEFAULT_FLAVOR_TEXT_NUMBER_OF_WORDS ='3 + d20.roll("1d6").total'

# fonts to display flavor titles in Excel properly
DEFAULT_EXCEL_FONT = 'Segoe UI Historic'
DEFAULT_EXCEL_FLAVOR_FONT_SIZE = 9

#############################################
# "Common" is just English. 
# Additional languages can be added; a .txt file with one word per line should be in the lorem_text_fantasy directory:

dictionary_languages = {
        'Classical' : 'latin.txt',
        'Common': 'english.txt', # uses just English; file is empty and does nothing but prevent bugs. :-)
        'Classical': 'latin.txt', 
        'Regional' : 'greek.txt', 
        'Ancient': 'akkadian.txt',
        'Dwarven' : 'runes.txt',
        'Elvish' : 'sindarin.txt',
        #'Akkadian': 'akkadian.txt',   # commented ones have no equivalence in ACKS tables for language.        
        #'Arabic': 'arabic.txt',       # ... adjust to taste, however.
        #'Armenian': 'armenian.txt',
        #'Chinese': 'chinese.txt',
        #'Cyrilic': 'cyrillic.txt',
        #'Georgian': 'georgian.txt',
        #'Gothic': 'gothic_latin.txt',
        #'Hebrew': 'hebrew.txt',
        #'Hindi': 'hindi.txt'
        #'Kanji':'kanji.txt',
        #'Korean':'korean.txt',
        # 'Classical': 'arabic.txt',
    }

font_languages = {
        'Classical' : DEFAULT_EXCEL_FONT,
        'Common': DEFAULT_EXCEL_FONT,
        'Classical': DEFAULT_EXCEL_FONT, 
        'Regional' : DEFAULT_EXCEL_FONT, 
        'Ancient': DEFAULT_EXCEL_FONT,
        'Dwarven' : 'Noto Sans Runic',
        'Elvish' : 'Tengwar Annatar',
        # 'Akkadian': DEFAULT_EXCEL_FONT,   
        #'Arabic': DEFAULT_EXCEL_FONT,       
        #'Armenian': DEFAULT_EXCEL_FONT,
        #'Chinese': DEFAULT_EXCEL_FONT,
        #'Cyrillic': DEFAULT_EXCEL_FONT,
        #'Georgian': DEFAULT_EXCEL_FONT,
        #'Gothic': DEFAULT_EXCEL_FONT,
        #'Hebrew': DEFAULT_EXCEL_FONT,
        #'Hindi': DEFAULT_EXCEL_FONT,
        #'Kanji': DEFAULT_EXCEL_FONT,
        #'Korean': DEFAULT_EXCEL_FONT,
    }
lang_no_spaces = ['Chinese','Kanji','Korean']
lang_limit_40_chars = ['Akkadian','Ancient','Gothic']

##################### End of user-settable variables ###########################

# general

global list_of_words_to_not_capitalize
global complexity_table_list

# names
global list_of_names_tables_male, list_of_names_tables_female
global author_title_table, epithets_tables

# names of male and female
global complete_table_male_names, name_tables_male
global complete_table_female_names, name_tables_female

# book titles

global titles_adjective_1_list, titles_communication, titles_conjunction_about, titles_conjunction_by, titles_fixed
global titles_history_of, titles_negative_subject
global titles_noun_1_list, titles_noun_2_list, titles_person_1, titles_person_2
global titles_places_cities, titles_places_nations, titles_religious_starter
global titles_study_in_list, titles_study_of_list, titles_study_on_list, titles_study_verbing, titles_the_1 
global titles_template_list_general, titles_template_list_history, titles_template_list_occult, titles_template_list_theology 

list_of_words_to_not_capitalize = [
    ('The','the'),
    ('Of','of'),
    ('De','de'),
    ("D'","d'"),
]
complexity_table_list = ['BookComplexityForScope1','BookComplexityForScope2','BookComplexityForScope3','BookComplexityForScope4']

surnames_tables = {}
name_tables_male = {}
name_tables_female = {}

# title of author
author_title_table = r.RPG_table('_titles_person')

#epithets of author
epithets_table = r.RPG_table('_epithets')

# saints names
titles_saints_male=r.RPG_table('_names_saints_male')
titles_saints_female=r.RPG_table('_names_saints_female')
titles_saints_amalgamated= titles_saints_male + titles_saints_female

# famous names
titles_person_famous_male=r.RPG_table('_names_famous_male')
titles_person_famous_female=r.RPG_table('_names_famous_female')
titles_person_famous_amalgamated=titles_person_famous_male + titles_person_famous_female

# blank lists for later use
complete_table_male_names = r.RPG_table('_names_empty')
complete_table_male_names.description = "Male Names Amalgamated"
complete_table_female_names = r.RPG_table('_names_empty')
complete_table_female_names.description = "Female Names Amalgamated"

# book title fragments
titles_adjective_1_list = r.RPG_table('_book_titles_adjective_1')
titles_biography_starter=r.RPG_table('_book_titles_biography_starter')
titles_communication=r.RPG_table('_book_titles_communication')
titles_conjunction_about = r.RPG_table('_book_titles_conjunction_about')
titles_conjunction_by = r.RPG_table('_book_titles_conjunction_by')
titles_fixed = r.RPG_table('_book_titles_fixed')
titles_history_of = r.RPG_table('_book_titles_history')
titles_negative_subject = r.RPG_table('_book_titles_negative_subject')
titles_noun_1_list = r.RPG_table('_book_titles_noun_1')
titles_noun_2_list = r.RPG_table('_book_titles_noun_2')
titles_person_evil=r.RPG_table('_names_famous_evil')
titles_places_cities = r.RPG_table('_book_titles_places_cities')
titles_places_nations = r.RPG_table('_book_titles_places_nations')
titles_religious_starter = r.RPG_table('_book_titles_religious_starter')
titles_study_in_list = r.RPG_table('_book_titles_study_in')
titles_study_of_list = r.RPG_table('_book_titles_study_of')
titles_study_on_list = r.RPG_table('_book_titles_study_on')
titles_study_verbing = r.RPG_table('_book_titles_study_verbing')
titles_the_1 = r.RPG_table('_books_titles_the_1')

# book title templates

titles_template_list_general = r.RPG_table('_book_titles_templates_general')
titles_template_list_history = r.RPG_table('_book_titles_templates_history')
titles_template_list_occult = r.RPG_table('_book_titles_templates_occult')
titles_template_list_theology = r.RPG_table('_book_titles_templates_theology')

# name tables load

list_of_names_tables_male = [
        '_names_anglo_saxon_male',
        '_names_arabic_male',
        '_names_english_male',
        '_names_famous_male', 
        '_names_french_male', 
        '_names_norse_male',
        '_names_roman_male', 
        ]

list_of_names_tables_female = [
        '_names_arabic_female',
        '_names_anglo_saxon_female', 
        '_names_english_female',
        '_names_famous_female', 
        '_names_french_female', 
        '_names_norse_female',
        '_names_roman_female', 
        ]

list_of_surnames_tables = [
        ('_names_arabic_surnames'),
        ('_names_anglo_saxon_surnames'), 
        ('_names_english_surnames'),
        ('_names_famous_surnames'), 
        ('_names_french_surnames'), 
        ('_names_norse_surnames_female'),
        ('_names_norse_surnames_male'),
        ('_names_roman_surnames'),
        ]

# load name table dictionaries
for i in list_of_names_tables_male:
    name_tables_male[i] = r.RPG_table(i)
    complete_table_male_names = (name_tables_male[i]) + complete_table_male_names

for i in list_of_names_tables_female:
    name_tables_female[i] = r.RPG_table(i)
    complete_table_female_names = (name_tables_female[i]) + complete_table_female_names

    # surnames
for i in list_of_surnames_tables:
    surnames_tables[i] = r.RPG_table(i) # creates dictionary containing a table for each nationality.

######################## FUNCTIONS ########################

def book_characteristics(books):
    book_attributes = [attribute for attribute in dir(books[1])
                   if not attribute.startswith('__')
                   and not callable(getattr(books[1], attribute))
                   ]
    # edit these to make appear in the Excel output in a different order
    book_variables_in_chosen_order = [
                                    'format',
                                    'materials',
                                    'number_volumes',
                                    'number_pages',
                                    'weight',
                                    'current_language',
                                    'original_language',
                                    'topic',
                                    'topic_apparent',
                                    'scope',
                                    'scope_esoteric',
                                    'complexity',
                                    'complexity_esoteric',
                                    'fraction_complete',
                                    'market_value',
                                    'book_title',
                                    'author_full',
                                    'translator_full_name',                             
                                    'book_title_flavor',
                                    'reading_time',
                                    'reference_time',
                                    'age_at_discovery',
                                    'number_extant_copies',
                                    'number_extant_available_to_place',
                                    'rarity_modifier',
                                    'libraries_it_is_in',
                                    'author_epithet',
                                    'author_name',
                                    'author_nationality',
                                    'author_title',
                                    'author_sex',
                                    'translator_name',
                                    'translator_nationality',
                                    'translator_sex',
                                    'translator_title',
                                    'weight_per_page',
                                    'template',
                                    'topic_title_form',
                                    'cost_per_page',
                                    'production_value',
                                    'literary_value_base',
                                    'literary_value_modified',
                                    'esoteric_literary_value_base',
                                    'esoteric_literary_value_modified',
                                    'year_discovered',
                                    'year_written',
                                    'is_a_translation',
                                    'book_type',
                                    ]
    
    # this bit adds any variables that have been omitted from the above list, so all will be displayed even if user error.
    for item in book_attributes:
        if item not in book_variables_in_chosen_order:
            book_variables_in_chosen_order.append(item)

    current_language_index = book_variables_in_chosen_order.index('current_language')
    flavor_title_index = book_variables_in_chosen_order.index('book_title_flavor')

    return book_variables_in_chosen_order, current_language_index+1, flavor_title_index+1 # index starts 0, Excel starts 1

def create_fantasy_book(book_type=None, **kwargs):
    ''' Returns a book object. Type can be default (normal), esoteric, authority, or magic'''
    book_type = string.capwords(str(book_type))
    return FantasyBook(**kwargs)

def export_books_to_excel (books,filename = 'books_spreadsheet_out.xlsx', worksheet = 'Book Hoard'):
    
    '''
    Takes a list of books and exports to Excel file. Default file 'books_spreadsheet_out.xlsx' in same folder, worksheet defaults to "Book Hoard."
    filename = 
    worksheet = 

    Can be passed in.

    If file and worksheet exist, is appended to. If file and/or worksheet do not exist, they are created.

    Note that export cannot function if the desired file is open. If this happens an option will be presented allowing user to close the Excel file and retry the save.
    '''

    book_columns,current_language_index, flavor_title_index = book_characteristics(books)
    
    try:
        wb = load_workbook(filename= filename)
    except:
        wb = Workbook()

    if worksheet in wb.sheetnames: 
        ws = wb[worksheet]
    else:
        ws = wb.create_sheet(title=worksheet)

    # column headers
    the_counter = 0
    for item in book_columns:
        the_counter += 1
        ws.cell(row=1,column=the_counter,value=item)
        ws.cell(row=1,column=the_counter).font = openpyxl_font(bold='bold',size=9)

    # Make sure can save

    try_to_save = True
    while try_to_save:
        try:
            wb.save(filename) 
        except:
            print ("You've probably got the excel file open; can't save.")
            user_response = input ("(T)ry again or (Q)uit? ")
            
            if user_response == "Q" or user_response == "q":
                try_to_save = False
                print ("Quitting without saving to Excel.")
                sys.exit()
                break
        else:
            try_to_save = False # ie succeeded

    # each row for a book
    the_counter = 0
    for book in books:
        row = []
        for attribute in book_columns:
            row.append(getattr(books[book],attribute))
        ws.append(row)
        the_counter += 1
        print ("Saving Book #" + str(the_counter) + "/" + str(len(books)) + " (" + str((int(100*the_counter/len(books)))) + "%)",end='\r')

            # now get language of the last row (just added) and set the proper font for the flavor title cell
        the_lang = ws.cell(row=ws.max_row,column=current_language_index)
        the_flavor = ws.cell(row=ws.max_row, column=flavor_title_index)
        the_flavor.font = openpyxl_font(name=font_languages[the_lang.value],size=DEFAULT_EXCEL_FLAVOR_FONT_SIZE)

    wb.save(filename)
    wb.close()
    print ('') # get off the same line
    print ("Exported to Excell file " + filename)

def import_language_words():
    ''' creates a dictionary of lists of various languages/character sets for the 'flavor text' titles of books based on their language.
        titles are generated with a lorem_ipsum algorithm from random words in *.txt files in the folder lorem_ipsum_fantasy.
        This is called just once as the program starts, and then the lists are passed to the lorem_ipsum_fantasy package.

        More languages and the like can be added in the dictionary_languages dictionary at the beginning of the program with the other constants. Key is the language; value is the name of the text file.
    '''
    ROOT_DIR = os.getcwd() # os.path.realpath(os.path.join(os.path.dirname(__file__), '..'))
    THIS_FOLDER = os.path.join((ROOT_DIR), 'lorem_text_fantasy')
    vocab_dictionary = {}

    for language,file in dictionary_languages.items():
        TARGET_LANGUAGE_FILE = os.path.join(THIS_FOLDER, file)
        the_words_imported = []
        with open(TARGET_LANGUAGE_FILE,encoding = 'utf8', mode='r') as f:
            for line in f.readlines():
                the_words_imported.append(line.strip())
            vocab_dictionary[language] = the_words_imported

    return vocab_dictionary

def produce_book_hoard (value=0,overshoot=False, **kwargs):
    ''' 
    produces a list of books worth the passed value. If overshoot is False, keeps total worth equal to or under value. If overshoot is true, then will produce a list that is _at least_ the passed value.

    Randomized characteristics unless keyword parameters are passed in. Those not passed with be randomized as far as it able (some values are interrelated, and so this can result in some slight deviations from the tables.)
    '''
    def update_book_status(the_count, running_total,value):
        print (" " * 80,end='\r') # blank the line
        print("Generating Book #" + str(the_count) + " --> " + str(running_total) + " gp/" + str (value) + " (" + str((int(100*running_total/value))) + "%)", end ='\r')

    books = {}

    while books == {}:
    
        running_total = 0
        the_count = 0

        while running_total < value:
            the_count += 1
            books[the_count] = create_fantasy_book(**kwargs)
            running_total += books[the_count].market_value
            update_book_status(the_count, running_total,value)

        if overshoot: 
            pass    
        else:
            running_total -= books[len(books)].market_value # subtract last value that put us over the top
            books.popitem() # delete last book which put over the top
            update_book_status(the_count, running_total,value)

        if books == {}:
            print ("Zero books made in hoard; retrying ....") # need better error checking to avoid endless loop if value too low.
    print ('') # get off the same line
    return books, running_total

def produce_number_of_books (number=1, **kwargs):
    ''' 
    Produces a given number of books. Randomized characteristics unless keyword parameters are passed in. Those not passed with be randomized as far as it able (some values are interrelated, and so this can result in some slight deviations from the tables.)
    '''

    def update_book_status(the_count, running_total,number):
        print (" " * 80,end='\r') # blank the line
        print("Generating Book #" + str(the_count) + "/" + str (number) + " (" + str((int(100*the_count/number))) + "%)" + " --> " + str(running_total) + " total gp value", end ='\r')

    books = {}

    while books == {}:
    
        running_total = 0
        the_count = 0

        while the_count < number:
            the_count += 1
            books[the_count] = create_fantasy_book(**kwargs)
            running_total += books[the_count].market_value
            update_book_status(the_count, running_total,number)

    print ('') # get off the same line
    return books, running_total
######################## CLASSES ########################

vocab_dictionary = import_language_words() # this is here because must come after definition of function

class FantasyBook():
    ''' Fantasy book object.'''

    # remember to add any variables added here to the self.XXXX list below, AND to the routine randomize_book_statistics 
    # if the value is to be set from a random table.

    def __init__(self,
        book_type = "Standard",         
        topic = "",
        topic_apparent = "",
        topic_title_form = "",
        book_title = "",
        book_title_flavor = '',
        author_sex = "",
        author_name = "",
        author_title = "",
        author_epithet = "",
        author_full = "",
        author_nationality = "",
        current_language = "",
        original_language = "",
        is_a_translation = False,
        translator_name = "",
        translator_nationality = "",
        translator_sex = '',
        translator_title = '',
        translator_full_name = '',
        format = "",
        materials = "",
        libraries_it_is_in = '',
        number_extant_copies = 0,
        number_extant_available_to_place = 0,
        scope = 0,
        scope_esoteric = 0,
        complexity = 0,
        complexity_esoteric = 0,
        age_at_discovery = 0,
        number_pages = 0,
        reading_time = 0,
        reference_time = 0,
        production_value = 0,
        literary_value_base = 0,
        literary_value_modified = 0,
        esoteric_literary_value_base = 0,
        esoteric_literary_value_modified = 0,
        rarity_modifier = 0,
        weight = 0,
        number_volumes = 0,
        year_discovered = 0,
        year_written = 0,
        market_value = 0,
        weight_per_page = 0,
        fraction_complete = 0,
        ):

        # set all values to whatever they were passed in 
        self.book_type = book_type
        self.topic = topic
        self.topic_apparent = topic_apparent
        self.topic_title_form = topic_title_form
        self.book_title = book_title
        self.book_title_flavor = book_title_flavor
        self.author_sex = author_sex
        self.author_name = author_name
        self.author_title = author_title
        self.author_epithet = author_epithet
        self.author_full = author_full
        self.author_nationality = author_nationality
        self.current_language = current_language
        self.original_language = original_language
        self.is_a_translation = is_a_translation
        self.translator_name = translator_name
        self.translator_nationality = translator_nationality
        self.translator_sex = translator_sex
        self.translator_title = translator_title
        self.translator_full_name = translator_full_name
        self.format = format
        self.materials = materials
        self.libraries_it_is_in = libraries_it_is_in
        self.number_extant_copies = number_extant_copies
        self.number_extant_available_to_place = number_extant_available_to_place
        self.scope = scope
        self.scope_esoteric = scope_esoteric
        self.complexity = complexity
        self.complexity_esoteric = complexity_esoteric
        self.age_at_discovery = age_at_discovery
        self.number_pages = number_pages
        self.reading_time = reading_time
        self.reference_time = reference_time
        self.production_value = production_value
        self.literary_value_base = literary_value_base
        self.literary_value_modified = literary_value_modified
        self.esoteric_literary_value_base = esoteric_literary_value_base
        self.esoteric_literary_value_modified = esoteric_literary_value_modified
        self.rarity_modifier = rarity_modifier
        self.weight = weight
        self.number_volumes = number_volumes
        self.year_discovered = year_discovered
        self.year_written = year_written
        self.market_value = market_value
        self.weight_per_page = weight_per_page
        self.fraction_complete = fraction_complete

        self.scope_set(self.scope)
        self.current_language_set(self.current_language)
        self.age_set(self.age_at_discovery)
        
        # translator_set must be called before original_language_set
        self.translator_set(translator_name=self.translator_name,translator_sex = self.translator_sex,translator_nationality=self.translator_nationality, translator_title=self.translator_title,translator_full_name=self.translator_full_name) 
        
        self.original_language_set(original_language = self.original_language, is_a_translation= self.is_a_translation)
        self.topic_set(topic = self.topic)
        self.topic_title_set(topic_title_form = self.topic_title_form)
        self.sex_set(author_sex = self.author_sex)
        self.author_epithet_set (author_epithet = self.author_epithet)
        self.author_name_set(author_name = self.author_name)
        self.author_title_set (author_title = self.author_title)
        self.author_full_set (author_full = self.author_full)
        self.complexity_set(complexity = self.complexity)
        self.format_set(format = self.format)
        self.book_title_set(book_title = self.book_title)
        self.materials_set(materials = self.materials)
        self.rarity_set(rarity_modifier = self.rarity_modifier, number_extant_copies = self.number_extant_copies, number_extant_available_to_place = self.number_extant_available_to_place)

        # these are all calculated based on other values above
        self.number_pages_set(number_pages = self.number_pages)
        self.reading_time_set(reading_time = self.reading_time)
        self.production_value_set(production_value = self.production_value)
        self.literary_value_set()
        self.weight_set(weight = self.weight)
        self.volumes_number_set(number_volumes = self.number_volumes)
        self.flavor_text_title_set(book_title_flavor = self.book_title_flavor)
        self.percentage_of_text_missing_set(fraction_complete = self.fraction_complete)
        self.esoteric_set()
        self.year_discovered = year_discovered
        self.year_written = year_written
        self.libraries_it_is_in = libraries_it_is_in

    def add(self,library):
        ''' Add this book to a given library'''
        self.libraries_it_is_in.append(library)

    def age_set(self,age=None):
        if not age:
            table_name = "BookAge_" + self.current_language # Ancient, Dwarvish, Elvish, Classical, Common are options
            dice_string = self.book_details_result_from_tables(table_name)
            if self.is_a_translation == True: 
                self.age_at_discovery = d20.roll(TRANSLATION_ADDITIONAL_AGE_OF_ORIGINAL).total # bonus to age if is translation.
            
            self.age_at_discovery = self.age_at_discovery + d20.roll(dice_string).total
        else:
            self.age_at_discovery = age
    
    def author_epithet_set (self, author_epithet=None):
        if not author_epithet:
            if CHANCE_OF_EPITHET_IN_AUTHOR_NAME > d20.roll("1d100").total:
                author_epithet = epithets_table.df.sample() # a random option is then chosen
                author_epithet = author_epithet.iloc[0,0]         
        self.author_epithet = author_epithet            

    def author_full_set (self, author_full=None):
        # put it all together
        if not author_full:

            if self.author_title != "" and self.author_title != "None": author_full = author_full.join([self.author_title," "])
            author_full += (self.author_name)
            if self.author_epithet != "" and self.author_epithet != "None": author_full = author_full + " " + self.author_epithet
        
        self.author_full = author_full
    
    def author_name_set(self,author_name=None,author_nationality = None):
        
        if not author_name:        
             
            author_name, author_nationality, _ = self.name_generate(sex = self.author_sex)
            
        self.author_name = author_name
        self.author_nationality = author_nationality

    def author_title_set(self, author_title=None):
        if not author_title:
            self.author_title = self.person_title_generate(sex = self.author_sex)
        else:
            self.author_title = author_title

    def book_details_result_from_tables(self,table_for_value,roll_result=None):
        ''' Checks table aaDiceTypeToRoll to see what dice to roll, and then rolls and checks the result on the given table.
        All tables that use this should have:
        1) An entry in SQL table aaDiceTypeToRoll with the same table name that matches table to roll on.
        2) The SQL table to be rolled on should have only two columns: one called 'DieRange' and the other 'Result'. 

        If roll_result is passed in as an integer value, then the table looks up as if that roll had been made (i.e., not random).
        '''

        # what dice to roll on the table   
       
        if not roll_result:
            dice_string = r.what_dice_to_roll(table_for_value) # returns a list
            
            if dice_string == []:
                raise ValueError("Empty dice list returned -- does SQL table 'aaDiceTypeToRoll' have an entry for this table?")
            else:
                roll_result = r.d20.roll(dice_string).total 

        # actually pass the roll value now that we know what dice have given
        t = r.RPG_table(table_for_value)
        rolled_row = t.roll(roll_result)
        return rolled_row 
    
    def book_title_set(
            self,
            adjective_1=None, 
            noun_1=None, 
            noun_2=None, 
            study_of=None, 
            study_in=None, 
            study_on=None, 
            template=None,
            book_title=None,
            conjunction_about=None,
            conjunction_by=None,
            negative_1=None,
            place_city=None,
            place_nation=None,
            religious_starter=None,
            verbing=None,
            saint=None,
            person_1=None,
            person_2=None,
            person_famous=None,
            communication=None,
            biography_starter=None,
            person_evil=None,
            history_of=None,
            the_1=None,

            ):

        avoid_special_class_of_title = True

        if book_title: 
            self.book_title = book_title
            self.template = "Final title passed in as: " + book_title
        
        else:
            if not template: template = titles_template_list_general.df.sample().iloc[0,0]
            topic = self.topic_title_form

            if "theology" in topic.lower():
                while ("religious" not in template) and ("biography" not in template):
                    template = titles_template_list_theology.df.sample().iloc[0,0]

            if "history" in topic.lower():
                while ("history" not in template) and ("biography" not in template):
                    template = titles_template_list_history.df.sample().iloc[0,0]

            if ("occult" in topic.lower()) or ("apostasy" in topic.lower()) or ("black lore" in topic.lower()):
                while ("occult" not in template) and ("negative" not in template) and {"evil" not in template} and ("biography" not in template):
                    template = titles_template_list_occult.df.sample().iloc[0,0]
            
            # refactor this code eventually into loop with eval()

            if not adjective_1 and "{adjective_1}" in template: adjective_1 = titles_adjective_1_list.df.sample().iloc[0,0]
            if not noun_1 and "{noun_1}" in template: noun_1 = titles_noun_1_list.df.sample().iloc[0,0]
            if not noun_2 and "{noun_2}" in template: noun_2 = titles_noun_2_list.df.sample().iloc[0,0]
            if not study_of and "{study_of}" in template: study_of = titles_study_of_list.df.sample().iloc[0,0]
            if not study_in and "{study_in}" in template: study_in = titles_study_in_list.df.sample().iloc[0,0]
            if not study_on and "{study_on}" in template: study_on = titles_study_on_list.df.sample().iloc[0,0]
            if not conjunction_about and "{conjunction_about}" in template: conjunction_about = titles_conjunction_about.df.sample().iloc[0,0]
            if not conjunction_by and "{conjunction_by}" in template: conjunction_by = titles_conjunction_by.df.sample().iloc[0,0]
            if not negative_1 and "{negative_1}" in template: negative_1 = titles_negative_subject.df.sample().iloc[0,0]
            if not place_city and "{place_city}" in template: place_city = titles_places_cities.df.sample().iloc[0,0]
            if not place_nation and "{place_nation}" in template: place_nation = titles_places_nations.df.sample().iloc[0,0]
            if not religious_starter and "{religious_starter}" in template: religious_starter = titles_religious_starter.df.sample().iloc[0,0]
            if not verbing and "{verbing}" in template: verbing = titles_study_verbing.df.sample().iloc[0,0]
            if not saint and "{saint}" in template: saint = titles_saints_amalgamated.df.sample().iloc[0,0]
            if not person_famous and "{person_famous}" in template: person_famous = titles_person_famous_amalgamated.df.sample().iloc[0,0]
            if not communication and "{communication}" in template: communication = titles_communication.df.sample().iloc[0,0]
            if not biography_starter and "{biography_starter}" in template: biography_starter = titles_biography_starter.df.sample().iloc[0,0]
            if not person_evil and "{person_evil}" in template: person_evil = titles_person_evil.df.sample().iloc[0,0]
            if not person_1 and "{person_1}" in template: person_1, _ , _ = self.name_generate() # 2nd,3rd are nation, sex which we don't need
            if not person_2 and "{person_2}" in template: person_2, _, _ = self.name_generate()
            if not history_of and "{history_of}" in template: history_of = titles_history_of.df.sample().iloc[0,0]
            if not the_1 and "{the_1}" in template: the_1 = titles_the_1.df.sample().iloc[0,0]
            
            self.template = template

            self.book_title = template.format(
                adjective_1 = adjective_1, 
                noun_1 = noun_1, 
                noun_2 = noun_2, 
                study_of = study_of, 
                study_in = study_in, 
                study_on = study_on, 
				topic = topic,
				conjunction_about=conjunction_about,
				conjunction_by=conjunction_by,
				negative_1=negative_1,
				place_city=place_city,
				place_nation=place_nation,
				religious_starter=religious_starter,
				verbing=verbing,
				saint=saint,
				person_1=person_1,
				person_2=person_2,
				person_famous=person_famous,
				communication=communication,
				biography_starter=biography_starter,
				person_evil=person_evil,
				history_of=history_of,
                the_1 = the_1
            )

    def complexity_set(self,complexity=None):
        if not complexity:
            complexity_from_table = self.book_details_result_from_tables(complexity_table_list[self.scope-1]) # Minus 1 since list index starts at zero.
            if complexity_from_table >= 1: complexity_from_table = int(complexity_from_table) # doesn't integerize 0.75
            self.complexity = complexity_from_table
        
        else:
            self.complexity = complexity   
    
    def current_language_set(self, current_language = None):
        if not current_language:
            self.current_language = self.book_details_result_from_tables("BookCurrentLanguage")
        else:
            self.current_language = current_language

    def esoteric_set(self):
        if self.topic != self.topic_apparent:
            esoteric_complexity_from_table = 0
            esoteric_ratios_correct = False

            while not esoteric_ratios_correct:
                # esoteric scope
                self.scope_esoteric = self.book_details_result_from_tables("BookScope")
                
                # esoteric complexity
                esoteric_complexity_from_table = self.book_details_result_from_tables(complexity_table_list[int(self.scope_esoteric)-1])
                if esoteric_complexity_from_table >= 1: esoteric_complexity_from_table = int(esoteric_complexity_from_table)      
                self.complexity_esoteric = esoteric_complexity_from_table

                # Is apparent ratio >= to the esoteric?
                if self.scope < 1: self.scope = 1 # self.scope can be very low if lots of the book is missing. Corrected below

                ratio_apparent = self.scope/self.complexity
                ratio_esoteric = self.scope_esoteric/self.complexity_esoteric

                if ratio_apparent >= ratio_esoteric: 
                    esoteric_ratios_correct = True
                    self.scope = self.scope * self.fraction_complete # restores from previous set = 1.
                    
                else:
                    pass
    
            self.esoteric_value_set()

    def esoteric_value_set (self):
        target_table = "BookLiteraryValueScope" + str(self.scope_esoteric)
        self.esoteric_literary_value_base = self.look_up_table(
            table_name=target_table,
            search_column="Complexity",
            search_term = self.complexity_esoteric,
            result_column="LiteraryValue"
            )
        self.esoteric_literary_value_modified = ceil (self.esoteric_literary_value_base * self.rarity_modifier) * self.number_pages * 5 # Writer of 18 Intelligence needed; hence the 5.
        self.market_value = ceil(self.literary_value_modified + self.esoteric_literary_value_modified + self.production_value)

    def format_set(self, format = None):
        if not format:
            target_table = "BookAge_Format_"
            if self.age_at_discovery < 11: target_table += "0001_0010"
            elif self.age_at_discovery < 51: target_table += "0011_0050"
            elif self.age_at_discovery < 101: target_table += "0051_0100"
            elif self.age_at_discovery < 501: target_table += "0101_0500"
            elif self.age_at_discovery < 1001: target_table += "0501_1000"
            elif self.age_at_discovery < 2001: target_table += "1001_2000"
            elif self.age_at_discovery < 10001: target_table += "2001_10000"

            self.format = self.book_details_result_from_tables(target_table)
        else:
            self.format = format
    
    def flavor_text_title_set(self, book_title_flavor=None):
        
        if not book_title_flavor:

            # Limit number chars (like Akkadian, gothic_latin)
            if  self.current_language in lang_limit_40_chars:
                limit_chars = 40 # These require only 40 chars or weird stuff happens. ? Unicode issue
            else:
                limit_chars = 0 # all the rest no limit

            # no spaces between:
            if self.current_language in lang_no_spaces:
                spaces = False # No spaces between works; Kanji looks better, for example.
            else:
                spaces = True # all the rest have spaces

            if self.current_language == "Common":
                book_title_flavor = self.book_title
            else:
                try:
                    
                    num_words_in_english_title = len(self.book_title.split())
                    num_words_in_flavor_title = eval(DEFAULT_FORMULA_CALC_NUM_FLAV_TEXT_WORDS_FROM_ORIG_TITLE)
                    if num_words_in_flavor_title <3: num_words_in_flavor_title = eval(DEFAULT_FLAVOR_TEXT_NUMBER_OF_WORDS)

                    book_title_flavor = str(
                        lf.words(vocab_dictionary[self.current_language],
                        count=num_words_in_flavor_title,
                        limit=limit_chars,
                        spaces = spaces)
                    )

                except:
                    book_title_flavor = "No flavor text designated for this language type."
                book_title_flavor = book_title_flavor.capitalize()

        self.book_title_flavor = book_title_flavor

    def literary_value_set (self):
        target_table = "BookLiteraryValueScope" + str(self.scope)
        self.literary_value_base = self.look_up_table(
            table_name=target_table,
            search_column="Complexity",
            search_term = self.complexity,
            result_column="LiteraryValue"
            )

        self.literary_value_modified = ceil(self.literary_value_base * self.rarity_modifier * self.number_pages)
        self.market_value = ceil(self.literary_value_modified + self.production_value)

    def look_up_table (self,result_column,table_name,search_column,search_term):
        query = 'SELECT {} from {} where {} LIKE "{}"'.format(result_column,table_name,search_column,search_term)
        t = r.LookUpTable(query = query)
        return t.result   
    
    def materials_set (self, materials=None):
        
        if not materials:
            target_table = "BookMaterials" + self.format
            self.materials = self.book_details_result_from_tables(target_table)

        else:
            self.materials = materials
   
    def name_generate(self,sex=None):
        # first name
        if sex == None:
            if d20.roll("1d100").total <  CHANCE_OF_FEMALE_AUTHOR: 
                sex = "Female"
            else: 
                sex = "Male"

        if sex == "Male": first_name = complete_table_male_names.df.sample()
        else: first_name = complete_table_female_names.df.sample()
        author_nationality = (first_name.iloc[0,1]) # the second column (i.e. index 1 since starts at 0) is the table for this type of name's surname.
                
        # surname
        last_name_table = (surnames_tables[author_nationality])
        last_name = last_name_table.df.sample()
        author_name = str(first_name.iloc[0,0]) + " " + str(last_name.iloc[0,0]) # first (0 index) item is the name

        return author_name, author_nationality, sex
    
    def number_pages_set(self, number_pages = None):
        if number_pages:
            self.number_pages = number_pages
            self.complexity = ceil((self.scope * 1000) / self.number_pages)

        else:
            self.number_pages = ceil((self.scope * 1000) / self.complexity) # note integer division // 

    def original_language_set(self, original_language=None,is_a_translation=False):
        
        if is_a_translation:
            self.is_a_translation = is_a_translation

        if (not is_a_translation) and (not original_language): 
            self.is_a_translation == False
            self.original_language == ''
            return

        if not original_language: # original language is empty
            original_language = self.book_details_result_from_tables("BookOriginalLanguage")

            while original_language == self.current_language:  
                original_language = self.book_details_result_from_tables("BookOriginalLanguage") # reroll until not the same
        
        self.original_language = original_language
        self.is_a_translation = True

    def percentage_of_text_missing_set(self,fraction_complete=None):
        
        if not fraction_complete:
            if CHANCE_OF_INCOMPLETE_WORK >= d20.roll("1d100").total:
                fraction_missing = round(d20.roll("1d99").total/100,2)
            else:
                fraction_missing = 0
            fraction_complete = 1 - fraction_missing
         
        self.scope = round(self.scope * 2.0 * fraction_complete) / 2.0 # the x2, then div 2 rounds to nearest 0.5
        if self.scope == 0: self.scope = 0.5

        self.reading_time = round(self.reading_time * 2.0 * fraction_complete) / 2.0
        self.reference_time = round(self.reference_time * 2.0 * fraction_complete) / 2.0

        self.number_pages = ceil(self.number_pages * fraction_complete)
        self.weight = ceil(self.weight * fraction_complete)
        self.market_value = ceil(self.market_value * fraction_complete)

        self.fraction_complete = round(fraction_complete,2)
    
    def person_title_generate (self,sex="Male"):
        global author_title_table
        author_title = ''

        if CHANCE_OF_TITLE_IN_AUTHOR_NAME >= d20.roll("1d100").total:

            author_title = str(author_title_table.df.sample().iloc[0,0])
           
        #  # male/female titles are separated by a slash in the SQL database  
            if author_title.__contains__("/"):
                title_split = author_title.split("/",2)
                
                if sex == "Male":
                    author_title = title_split[0]
                else:
                    author_title = title_split[1]
        
        return string.capwords(str(author_title))
    
    def production_value_set(self, production_value = None):

        if production_value:
            self.production_value = production_value
            self.cost_per_page = production_value / self.number_pages

        else:
            target_table = "BookProductionValue" + self.format
            self.cost_per_page = self.look_up_table(result_column="Cost",table_name=target_table,search_column="Material",search_term=self.materials)
            self.production_value = ceil(self.cost_per_page * self.number_pages)
        
    def rarity_set(self, rarity_modifier = None, number_extant_copies = None, number_extant_available_to_place = None):
        the_roll = d20.roll("1d100").total # this same value needed twice, so must roll it first so can be passed.

        if number_extant_copies:
            self.number_extant_copies = number_extant_copies
        
        else:
            dice_string_determine_number_copies = self.book_details_result_from_tables("BookRarityCopies", roll_result= the_roll)
            number_of_copies_roll = d20.roll(dice_string_determine_number_copies).total
            self.number_extant_copies = number_of_copies_roll
        
        if number_extant_available_to_place:
            self.number_extant_available_to_place = number_extant_available_to_place

        else:
            self.number_extant_available_to_place = (self.number_extant_copies - 1) # ie, less this one.

        if rarity_modifier:
            self.rarity_modifier = rarity_modifier
        
        else:
            self.rarity_modifier = self.book_details_result_from_tables("BookRarityModifier",roll_result=the_roll)

    def reading_time_set(self,reading_time = None):

        if reading_time:
            self.number_pages = READING_PAGES_PER_HOUR * reading_time
            self.reading_time = reading_time
            self.reference_time = reading_time

        else:
            self.reading_time = ceil(self.number_pages/READING_PAGES_PER_HOUR)
            self.reference_time = self.reading_time

    def refresh_number_pages(self):
            self.number_pages_set(number_pages = self.number_pages)
            self.reading_time_set()
            self.production_value_set()    

    def remove (self,library):
        ''' Remove this book from a given library'''
        try:
            self.libraries_it_is_in.remove(library)
        except ValueError:
            print ("The book _{}_ is not in {} library.".format(self.title, library))
    
    def scope_set(self, scope=None):
        if not scope:
            self.scope = self.book_details_result_from_tables("BookScope")
        else:
            self.scope = scope
    
    def sex_set (self, author_sex=None):
        if not author_sex:
            if d20.roll("1d100").total <= CHANCE_OF_FEMALE_AUTHOR: self.author_sex = "Female"
            else: self.author_sex = "Male"
        else:
            self.author_sex = author_sex
    
    def topic_set (self, topic=None):

        if not topic:
            topic = self.book_details_result_from_tables("BookTopicsACKS")
           
        self.topic = topic
        self.topic_apparent = topic     
        
        if "Esoteric" in self.topic:
                while "Esoteric" in topic: # keep picking apparent topic until not esoteric
                    topic = self.book_details_result_from_tables("BookTopicsACKS")
                    
                self.topic_apparent = topic
                  
    def topic_title_set(self,topic_title_form=None):
        if not topic_title_form:

            t = self.look_up_table(
                result_column="title_string",
                table_name= "_book_titles_topics", 
                search_column="Result",
                search_term = self.topic_apparent
                )
            
            t = t.split(";") # list is made by separating by semicolons
            t = random.choice(t) # a random option is then chosen
            self.topic_title_form = t
        
        else:
            self.topic_title_form = topic_title_form

    def translator_set (self,translator_name=None,translator_sex = None,translator_nationality=None, translator_title=None,translator_full_name=None):
        roll_to_see_if_it_is_a_translation = d20.roll("1d100").total

        self.is_a_translation = False
        self.translator_full_name = translator_full_name

        if (translator_name or translator_full_name) and (self.current_language not in ANCIENT_LANGUAGES_WHICH_WOULD_NOT_BE_TRANSLATED_INTO):
            self.is_a_translation = True
            self.translator_nationality = translator_nationality
            self.translator_title = translator_title
            self.translator_name = translator_name
            self.translator_sex = translator_sex

            if translator_full_name:
                self.translator_full_name = translator_full_name
            else:
                self.translator_full_name = self.translator_title + " " + self.translator_name

        elif (roll_to_see_if_it_is_a_translation < CHANCE_OF_BEING_TRANSLATION) and (self.current_language not in ANCIENT_LANGUAGES_WHICH_WOULD_NOT_BE_TRANSLATED_INTO):
            self.is_a_translation = True
            self.translator_name, self.translator_nationality, self.translator_sex = self.name_generate()
            self.translator_title = self.person_title_generate(sex = self.translator_sex)
            self.translator_full_name = self.translator_title + " " + self.translator_name

        else:
            
            self.is_a_translation = False

    def volumes_number_set(self, number_volumes = None):
        if number_volumes:
            self.number_volumes = number_volumes

            if self.format == "Codex":
                self.number_pages = ceil (self.number_volumes * PAGES_PER_VOLUME_FOR_CODEX)
                self.weight = ceil ((self.number_pages * self.weight_per_page) + (self.number_volumes * WEIGHT_PER_VOLUME_OF_CODEX))
            
            elif self.format == "Scroll":
                self.number_pages = ceil (self.number_volumes * PAGES_PER_VOLUME_FOR_SCROLL)
                self.weight = ceil((self.number_pages * self.weight_per_page) + (self.number_volumes * WEIGHT_PER_VOLUME_OF_SCROLL))

            elif self.format == "Tablet":
                self.number_volumes = 1 # ie, never multivolume
            
            else:
                raise ValueError("Format has a problem: is not a Codex, Scroll, or Tablet.")

            self.refresh_number_pages()

        else:

            if self.format == "Codex":
                self.number_volumes = ceil(self.number_pages/PAGES_PER_VOLUME_FOR_CODEX)
                self.weight = self.weight + (self.number_volumes * WEIGHT_PER_VOLUME_OF_CODEX)
            
            elif self.format == "Scroll":
                self.number_volumes = ceil(self.number_pages/PAGES_PER_VOLUME_FOR_SCROLL)
                self.weight = self.weight + (self.number_volumes * WEIGHT_PER_VOLUME_OF_SCROLL)
            
            elif self.format == "Tablet":
                self.number_volumes = 1 # ie, never multivolume

            else:
                raise ValueError("Format has a problem: is not a Codex, Scroll, or Tablet.")
        
    def weight_set(self,weight=None):
        self.weight_per_page = self.look_up_table(result_column="Result",table_name="BookWeight",search_column="Material",search_term=self.materials)

        if weight:
            self.weight = weight
            self.number_pages  = ceil(self.weight / self.weight_per_page)
            self.refresh_number_pages()

        else:
            self.weight = ceil(self.weight_per_page * self.number_pages)

class MagicBook(FantasyBook):
    ''' Subclass of fantasy book, that has a few extra values.'''
    def __init__ (self,
        book_type = "Magic",
    ):
        super().__init__(self)

######################## main() ########################

# books, books_value = produce_book_hoard(value=15000,overshoot=True,fraction_complete=0.1)
books, books_value = produce_number_of_books(number = 5,topic='Healing')

# print_book_hoard(books)
export_books_to_excel(books)

print ('TOTAL: ' + str(books_value))
print ('Number of books: ' + str (len(books)))